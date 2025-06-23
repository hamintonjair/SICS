# backend/app/controllers/actividad_controller.py
from flask import request, current_app, jsonify, send_file, g
from bson import ObjectId
from werkzeug.utils import secure_filename
from datetime import datetime, timezone
import os
import logging
import traceback
import tempfile  # Para manejo de archivos temporales
from bson.errors import InvalidId
from app.models.actividad import ActividadModel, actividad_schema, ActividadSchema

# Configurar el logger
logger = logging.getLogger(__name__)

def configurar_subida_logo():
    upload_folder = os.path.join(current_app.root_path, '..', '..', 'uploads', 'logos')
    os.makedirs(upload_folder, exist_ok=True)
    return upload_folder

def crear_actividad():
    try:
        from app.models.actividad import ActividadModel, actividad_schema, ActividadSchema
        
        # Obtener datos del formulario
        datos = request.get_json()
        if not datos:
            logger.error("No se recibieron datos en la solicitud")
            return jsonify({
                'success': False,
                'message': 'No se recibieron datos en la solicitud'
            }), 400
            
        logger.info(f"Datos recibidos para crear actividad: {datos}")
        
        # Validar que los datos sean un diccionario
        if not isinstance(datos, dict):
            logger.error(f"Los datos deben ser un diccionario, se recibió: {type(datos)}")
            return jsonify({
                'success': False,
                'message': 'Los datos deben ser un diccionario',
                'tipo_recibido': str(type(datos))
            }), 400
        
        # Validar que todos los campos requeridos estén presentes
        required_fields = [
            'tema', 'objetivo', 'lugar', 'dependencia', 'fecha',
            'hora_inicio', 'hora_fin', 'linea_trabajo_id'
        ]
        
        # Si no hay creado_por, usar un valor por defecto
        if 'creado_por' not in datos:
            datos['creado_por'] = 'sistema'  # O cualquier otro valor por defecto
            logger.info("No se proporcionó 'creado_por', usando valor por defecto")
        
        # Asegurarse de que funcionario_id esté presente, si no, usar creado_por
        if 'funcionario_id' not in datos:
            datos['funcionario_id'] = datos['creado_por']
            logger.info(f"Usando creado_por como funcionario_id: {datos['funcionario_id']}")
        
        # Verificar campos faltantes
        missing_fields = [field for field in required_fields if field not in datos]
        if missing_fields:
            logger.error(f"Faltan campos requeridos: {missing_fields}")
            return jsonify({
                'success': False,
                'message': 'Faltan campos requeridos',
                'missing_fields': missing_fields,
                'datos_recibidos': {k: v for k, v in datos.items() if k in required_fields}
            }), 422
            
        # Validar formato de fecha
        try:
            fecha_str = datos['fecha']
            if isinstance(fecha_str, str):
                # Si es string, convertirlo a datetime
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
                # Asegurarse de que la fecha esté en UTC
                if hasattr(fecha_obj, 'tzinfo') and fecha_obj.tzinfo is not None:
                    fecha_obj = fecha_obj.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                datos['fecha'] = fecha_obj
            elif isinstance(fecha_str, datetime):
                # Si ya es datetime, asegurarse de que esté en UTC
                if hasattr(fecha_str, 'tzinfo') and fecha_str.tzinfo is not None:
                    fecha_obj = fecha_str.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                else:
                    fecha_obj = fecha_str
                datos['fecha'] = fecha_obj
            else:
                raise ValueError(f"Formato de fecha no soportado: {type(fecha_str)}")
                
            logger.info(f"Fecha convertida: {datos['fecha']} ({type(datos['fecha']).__name__})")
        except Exception as e:
            logger.error(f"Error al procesar la fecha: {str(e)}")
            return jsonify({
                'success': False,
                'message': 'Error en el formato de fecha',
                'error': str(e)
            }), 400
            
        # Validar formato de horas
        try:
            datetime.strptime(datos['hora_inicio'], '%H:%M')
            datetime.strptime(datos['hora_fin'], '%H:%M')
        except (ValueError, TypeError) as e:
            logger.error(f"Error en formato de hora: {str(e)}")
            return jsonify({
                'success': False,
                'message': 'Formato de hora inválido. Use HH:MM',
                'error': str(e)
            }), 422
        
        # Usar el usuario del sistema como creador si no se proporciona
        if 'creado_por' not in datos:
            datos['creado_por'] = 'sistema'
                
        # Usar creado_por como funcionario_id si no se proporciona
        if 'funcionario_id' not in datos:
            datos['funcionario_id'] = datos['creado_por']
        
        # Validar los datos con el esquema
        try:
            # Asegurarse de que la fecha sea un string ISO 8601
            if 'fecha' in datos and isinstance(datos['fecha'], datetime):
                datos['fecha'] = datos['fecha'].isoformat()
                
            # Asegurarse de que los IDs sean strings
            for field in ['linea_trabajo_id', 'funcionario_id', 'creado_por']:
                if field in datos and datos[field]:
                    if isinstance(datos[field], ObjectId):
                        datos[field] = str(datos[field])
                    elif not isinstance(datos[field], str):
                        datos[field] = str(datos[field])
            
            logger.info(f"Datos antes de validar con esquema: {datos}")
            
            # Validar manualmente el esquema
            schema = ActividadSchema()
            errors = schema.validate(datos)
            
            if errors:
                logger.error(f"Errores de validación: {errors}")
                return jsonify({
                    'success': False,
                    'message': 'Error de validación de datos',
                    'errors': errors,
                    'datos_recibidos': datos
                }), 422
                
            # Si pasa la validación, cargar los datos
            result = schema.load(datos)
            logger.info(f"Datos validados correctamente: {result}")
        except Exception as e:
            error_details = {
                'error': str(e),
                'error_type': type(e).__name__,
                'validation_errors': getattr(e, 'messages', 'No hay mensajes de validación'),
                'datos_recibidos': datos
            }
            logger.error(f"Error de validación: {error_details}")
            return jsonify({
                'success': False,
                'message': 'Error de validación de datos',
                **error_details
            }), 422
        
        # Crear la actividad
        actividad_model = ActividadModel()
        actividad_id = actividad_model.crear_actividad(datos)
        
        if not actividad_id:
            raise Exception("No se pudo crear la actividad")
            
        return jsonify({
            'success': True,
            'message': 'Actividad creada exitosamente',
            'actividad_id': str(actividad_id)
        }), 201
        
    except Exception as e:
        logger.error(f"Error al crear actividad: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'Error al crear la actividad',
            'error': str(e),
            'error_type': type(e).__name__
        }), 500

def obtener_actividades():
    try:
        from app.models.actividad import ActividadModel
        
        # Obtener filtros de la URL
        filtros = request.args.to_dict()
        
        # Inicializar el modelo
        actividad_model = ActividadModel()
        
        # Obtener actividades con los filtros proporcionados
        actividades = actividad_model.obtener_actividades(filtros)
        
        return jsonify({
            'success': True,
            'data': actividades
        })
        
    except Exception as e:
        logger.error(f"Error al obtener actividades: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Error al obtener las actividades',
            'error': str(e)
        }), 500

def obtener_actividad(actividad_id):
    """
    Obtiene los detalles de una actividad específica por su ID
    """
    from flask import jsonify
    from bson import ObjectId
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Obteniendo detalles de la actividad {actividad_id}")
        
        # Validar el ID de la actividad
        if not ObjectId.is_valid(actividad_id):
            logger.error(f"ID de actividad inválido: {actividad_id}")
            return jsonify({
                'success': False,
                'message': 'ID de actividad inválido'
            }), 400
        
        # Importar modelos dentro del bloque try para evitar importaciones circulares
        from app.models.actividad import ActividadModel
        from app.models.beneficiario import BeneficiarioModel
        
        # Obtener la actividad
        actividad_model = ActividadModel()
        actividad = actividad_model.obtener_actividad_por_id(actividad_id)
        
        if not actividad:
            logger.warning(f"No se encontró la actividad con ID: {actividad_id}")
            return jsonify({
                'success': False,
                'message': 'Actividad no encontrada'
            }), 404
        
        # Obtener información detallada de los beneficiarios si hay asistentes
        if 'asistentes' in actividad and actividad['asistentes']:
            logger.info(f"Obteniendo información de {len(actividad['asistentes'])} asistentes")
            beneficiario_model = BeneficiarioModel()
            
            for asistente in actividad['asistentes']:
                try:
                    beneficiario_id = asistente.get('beneficiario_id')
                    if not beneficiario_id:
                        logger.warning("Asistente sin beneficiario_id")
                        continue
                        
                    beneficiario = beneficiario_model.obtener_beneficiario_por_id(beneficiario_id)
                    if beneficiario:
                        # Agregar información del beneficiario al asistente
                        asistente['beneficiario'] = {
                            'nombre_completo': beneficiario.get('nombre_completo', ''),
                            'tipo_documento': beneficiario.get('tipo_documento', ''),
                            'numero_documento': beneficiario.get('numero_documento', '')
                        }
                except Exception as e:
                    logger.error(f"Error al obtener información del beneficiario: {str(e)}")
                    continue
        
        # Retornar la actividad con la información de los asistentes
        return jsonify({
            'success': True,
            'data': actividad
        })
        
    except Exception as e:
        logger.error(f"Error al obtener la actividad: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'Error al obtener la actividad',
            'error': str(e)
        }), 500

def actualizar_actividad(actividad_id):
    """
    Actualiza una actividad existente
    """
    logger.info(f"[ACTUALIZAR_ACTIVIDAD] Iniciando actualización de actividad {actividad_id}")
    
    try:
        # Verificar si hay datos JSON
        if not request.is_json:
            logger.error("[ACTUALIZAR_ACTIVIDAD] No se recibieron datos JSON en la solicitud")
            return jsonify({
                'success': False,
                'message': 'Se esperaba un cuerpo JSON en la solicitud',
                'error': 'NO_JSON_DATA'
            }), 400
            
        # Obtener datos JSON
        try:
            datos = request.get_json()
            logger.info(f"[ACTUALIZAR_ACTIVIDAD] Datos recibidos: {datos}")
            
            # Validar que los datos sean un diccionario
            if not isinstance(datos, dict):
                logger.error(f"[ACTUALIZAR_ACTIVIDAD] Los datos deben ser un objeto JSON. Tipo recibido: {type(datos)}")
                return jsonify({
                    'success': False,
                    'message': 'Los datos deben ser un objeto JSON',
                    'error': 'INVALID_JSON_FORMAT'
                }), 400
                
        except Exception as e:
            error_msg = f"[ACTUALIZAR_ACTIVIDAD] Error al procesar los datos JSON: {str(e)}"
            logger.error(error_msg)
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Traceback: {traceback.format_exc()}")
            return jsonify({
                'success': False,
                'message': 'Error al procesar los datos de la solicitud',
                'error': 'INVALID_JSON',
                'details': str(e)
            }), 400
            
        if not datos:
            logger.error("[ACTUALIZAR_ACTIVIDAD] No se recibieron datos para actualizar")
            return jsonify({
                'success': False,
                'message': 'No se recibieron datos para actualizar',
                'error': 'EMPTY_DATA'
            }), 400
            
        # Validar que el ID sea un ObjectId válido
        try:
            if not ObjectId.is_valid(actividad_id):
                logger.error(f"[ACTUALIZAR_ACTIVIDAD] ID de actividad inválido: {actividad_id}")
                return jsonify({
                    'success': False,
                    'message': 'ID de actividad inválido',
                    'campo': 'id',
                    'error': 'INVALID_ACTIVITY_ID'
                }), 400
        except Exception as e:
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Error al validar el ID de actividad: {str(e)}")
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Traceback: {traceback.format_exc()}")
            return jsonify({
                'success': False,
                'message': 'Error al validar el ID de actividad',
                'error': 'VALIDATION_ERROR',
                'details': str(e)
            }), 500
    
        # Crear un nuevo diccionario con los campos permitidos
        datos_actualizados = {}
        
        # Lista de campos permitidos para actualización
        campos_permitidos = [
            'tema', 'objetivo', 'lugar', 'dependencia', 'fecha',
            'hora_inicio', 'hora_fin', 'linea_trabajo_id', 'estado',
            'asistentes', 'logo_url', 'actualizado_por', 'funcionario_id'
        ]
        
        # Copiar solo los campos permitidos que existen en los datos
        for campo in campos_permitidos:
            if campo in datos and datos[campo] is not None:  # Solo copiar si el campo existe y no es None
                datos_actualizados[campo] = datos[campo]
        
        # Si no hay campos para actualizar
        if not datos_actualizados:
            logger.warning("[ACTUALIZAR_ACTIVIDAD] No se proporcionaron campos válidos para actualizar")
            return jsonify({
                'success': True,
                'message': 'No se realizaron cambios (sin campos válidos para actualizar)',
                'modificados': 0
            })
        
        logger.info(f"[ACTUALIZAR_ACTIVIDAD] Datos a actualizar: {datos_actualizados}")
        
        # Validar campos requeridos solo si se están actualizando
        campos_requeridos = {
            'tema': 'El tema es requerido',
            'objetivo': 'El objetivo es requerido',
            'lugar': 'El lugar es requerido',
            'dependencia': 'La dependencia es requerida',
            'fecha': 'La fecha es requerida',
            'hora_inicio': 'La hora de inicio es requerida',
            'hora_fin': 'La hora de fin es requerida',
            'linea_trabajo_id': 'La línea de trabajo es requerida',
            'funcionario_id': 'El funcionario es requerido'
        }
        
        # Validar solo los campos que se están actualizando
        for campo, mensaje in campos_requeridos.items():
            if campo in datos_actualizados:  # Solo validar campos que se están actualizando
                if not datos_actualizados[campo]:
                    logger.error(f"[ACTUALIZAR_ACTIVIDAD] Campo requerido vacío: {campo}")
                    return jsonify({
                        'success': False,
                        'message': mensaje,
                        'campo': campo,
                        'error': 'EMPTY_REQUIRED_FIELD'
                    }), 400
        
        # Inicializar el modelo de actividad
        try:
            from app.models.actividad import ActividadModel
            actividad_model = ActividadModel()
            logger.info("[ACTUALIZAR_ACTIVIDAD] Modelo de actividad inicializado correctamente")
            
            # Verificar si la actividad existe
            logger.info(f"[ACTUALIZAR_ACTIVIDAD] Buscando actividad con ID: {actividad_id}")
            actividad_existente = actividad_model.obtener_actividad_por_id(actividad_id)
            
            if not actividad_existente:
                logger.error(f"[ACTUALIZAR_ACTIVIDAD] No se encontró la actividad con ID: {actividad_id}")
                return jsonify({
                    'success': False,
                    'message': 'No se encontró la actividad especificada',
                    'error': 'ACTIVITY_NOT_FOUND'
                }), 404
                
            # Agregar fecha de actualización
            datos_actualizados['fecha_actualizacion'] = datetime.utcnow()
            logger.info(f"[ACTUALIZAR_ACTIVIDAD] Fecha de actualización: {datos_actualizados['fecha_actualizacion']}")
            
            # Actualizar la actividad
            logger.info("[ACTUALIZAR_ACTIVIDAD] Actualizando actividad...")
            resultado = actividad_model.actualizar_actividad(actividad_id, datos_actualizados)
            
            if not resultado or not hasattr(resultado, 'modified_count') or resultado.modified_count == 0:
                logger.warning("[ACTUALIZAR_ACTIVIDAD] No se realizaron cambios en la actividad")
                return jsonify({
                    'success': True,
                    'message': 'No se realizaron cambios en la actividad',
                    'modificados': 0
                })
                
            logger.info(f"[ACTUALIZAR_ACTIVIDAD] Actividad {actividad_id} actualizada exitosamente")
            return jsonify({
                'success': True,
                'message': 'Actividad actualizada exitosamente',
                'modificados': resultado.modified_count
            })
            
        except Exception as e:
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Error al actualizar la actividad en la base de datos: {str(e)}")
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Traceback: {traceback.format_exc()}")
            return jsonify({
                'success': False,
                'message': 'Error al actualizar la actividad en la base de datos',
                'error': 'DATABASE_ERROR',
                'details': str(e)
            }), 500
            
    except Exception as e:
        logger.error(f"[ACTUALIZAR_ACTIVIDAD] Error inesperado al actualizar la actividad: {str(e)}")
        logger.error(f"[ACTUALIZAR_ACTIVIDAD] Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': 'Error inesperado al actualizar la actividad',
            'error': 'INTERNAL_SERVER_ERROR',
            'details': str(e)
        }), 500

def eliminar_actividad(actividad_id):
    try:
        from app.models.actividad import ActividadModel
        
        actividad_model = ActividadModel()
        eliminados = actividad_model.eliminar_actividad(actividad_id)
        
        if eliminados == 0:
            return jsonify({
                'success': False,
                'message': 'No se encontró la actividad'
            }), 404
            
        return jsonify({
            'success': True,
            'message': 'Actividad eliminada exitosamente',
            'eliminados': eliminados
        })
        
    except Exception as e:
        logger.error(f"Error al eliminar la actividad: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Error al eliminar la actividad',
            'error': str(e)
        }), 500

def registrar_asistencias(actividad_id):
    try:
        from app.models.actividad import ActividadModel
        from app.routes.auth import get_jwt_identity
        
        usuario_actual = get_jwt_identity()
        datos = request.get_json()
        
        if 'asistentes' not in datos:
            return jsonify({
                'success': False,
                'message': 'Se requiere la lista de asistentes'
            }), 400
            
        # Agregar información de actualización
        datos['actualizado_por'] = usuario_actual['id']
        
        actividad_model = ActividadModel()
        modificados = actividad_model.registrar_asistencias(actividad_id, datos['asistentes'])
        
        return jsonify({
            'success': True,
            'message': 'Asistencias registradas exitosamente',
            'modificados': modificados
        })
        
    except Exception as e:
        logger.error(f"Error al registrar asistencias: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Error al registrar asistencias',
            'error': str(e)
        }), 500
        if not actividad:
            logger.error(f"Actividad no encontrada: {actividad_id}")
            return jsonify({
                'success': False,
                'message': 'Actividad no encontrada'
            }), 404
            
        logger.info(f"Actividad encontrada: {actividad.get('nombre', 'Sin nombre')}")
        
        # Crear un libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Encabezados
        headers = [
            'N°', 'FECHA DE REGISTRO', 'NOMBRE', 'TIPO DOCUMENTO', 'IDENTIFICACIÓN', 
            'GÉNERO', 'EDAD', 'COMUNA', 'BARRIO', 'TELÉFONO', 'CORREO ELECTRÓNICO', 
            'ESTUDIA ACTUALMENTE', 'SABE LEER', 'SABE ESCRIBIR', 'TIPO DE VIVIENDA', 
            'SITUACIÓN LABORAL', 'GRUPO ÉTNICO', 'AYUDA HUMANITARIA', 'DISCAPACIDAD', 
            'TIPO DE DISCAPACIDAD', 'NOMBRE DE LA CUIDADORA', 'LABORA ACTUALMENTE', 
            'VÍCTIMA DE CONFLICTO'
        ]
        
        # Agregar encabezados de columnas
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        # Procesar asistentes (comenzando después de los encabezados)
        row_num = start_row + 1
        for i, asistente in enumerate(actividad.get('asistentes', []), 1):
            try:
                beneficiario = asistente.get('beneficiario', {})
                
                # Procesar fecha de registro
                fecha_registro = asistente.get('fecha_asistencia', '')
                if fecha_registro and isinstance(fecha_registro, str):
                    try:
                        fecha_obj = datetime.strptime(fecha_registro, '%Y-%m-%dT%H:%M:%S.%fZ')
                        fecha_registro = fecha_obj.strftime('%d/%m/%Y %H:%M')
                    except (ValueError, TypeError):
                        fecha_registro = fecha_registro.split('T')[0]
                
                # Obtener edad exacta
                edad = str(beneficiario.get('edad', '')) if str(beneficiario.get('edad', '')).isdigit() else ''
                
                # Preparar fila
                row = [
                    i,  # N°
                    fecha_registro,  # FECHA DE REGISTRO
                    beneficiario.get('nombre_completo', ''),  # NOMBRE
                    beneficiario.get('tipo_documento', ''),  # TIPO DOCUMENTO
                    str(beneficiario.get('numero_documento', '')),  # IDENTIFICACIÓN
                    beneficiario.get('genero', ''),  # GÉNERO
                    edad,  # EDAD (número exacto)
                    beneficiario.get('comuna', ''),  # COMUNA
                    beneficiario.get('barrio', ''),  # BARRIO
                    beneficiario.get('numero_celular', ''),  # TELÉFONO
                    beneficiario.get('correo_electronico', ''),  # CORREO ELECTRÓNICO
                    'Sí' if beneficiario.get('estudia_actualmente', False) else 'No',  # ESTUDIA ACTUALMENTE
                    beneficiario.get('nivel_educativo', ''),  # NIVEL EDUCATIVO
                    'Sí' if beneficiario.get('sabe_leer', False) else 'No',  # SABE LEER
                    'Sí' if beneficiario.get('sabe_escribir', False) else 'No',  # SABE ESCRIBIR
                    beneficiario.get('tipo_vivienda', ''),  # TIPO DE VIVIENDA
                    beneficiario.get('situacion_laboral', ''),  # SITUACIÓN LABORAL
                    beneficiario.get('etnia', ''),  # GRUPO ÉTNICO
                    'Sí' if beneficiario.get('recibe_ayuda_humanitaria', False) else 'No',  # AYUDA HUMANITARIA
                    'Sí' if beneficiario.get('tiene_discapacidad', False) else 'No',  # DISCAPACIDAD
                    beneficiario.get('tipo_discapacidad', ''),  # TIPO DE DISCAPACIDAD
                    beneficiario.get('nombre_cuidador', ''),  # NOMBRE DE LA CUIDADORA
                    'Sí' if beneficiario.get('labora_actualmente', False) else 'No',  # LABORA ACTUALMENTE
                    'Sí' if beneficiario.get('victima_conflicto', False) else 'No'  # VÍCTIMA DE CONFLICTO
                ]
                
                # Agregar fila a la hoja
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                row_num += 1
                
            except Exception as e:
                logger.error(f"Error al procesar asistente {i}: {str(e)}")
                continue
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 50)  # Máximo 50 caracteres
        
        # Crear un archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            excel_path = tmp_file.name
        
        # Guardar el libro de Excel
        wb.save(excel_path)
        logger.info(f"Archivo Excel generado: {excel_path}")
        
        # Enviar el archivo
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=f"asistencia-{actividad.get('nombre', 'actividad')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error al exportar actividad {actividad_id}", exc_info=True)
        logger.error(f"Tipo de error: {type(e).__name__}")
        logger.error(f"Mensaje de error: {str(e)}")
        
        # Si hay un error, asegurarse de eliminar el archivo temporal
        if 'excel_path' in locals() and os.path.exists(excel_path):
            try:
                os.unlink(excel_path)
                logger.info(f"Archivo temporal eliminado: {excel_path}")
            except Exception as e2:
                logger.error(f"Error al eliminar archivo temporal {excel_path}: {str(e2)}")
        
        return jsonify({
            'success': False,
            'message': f'Error al exportar la actividad: {str(e)}',
            'error_type': type(e).__name__,
            'error_details': str(e)
        }), 500

    try:
        from app.models.actividad import ActividadModel
        from app.models.beneficiario import BeneficiarioModel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import numbers
        import tempfile
        import os
        import requests
        from io import BytesIO
        from bson import ObjectId
        from datetime import datetime
        from PIL import Image as PILImage
        from PIL import ImageDraw, ImageFont
        
        logger.info(f"Iniciando exportación de actividad a Excel: {actividad_id}")
        logger.info(f"Columnas seleccionadas: {columnas}")
        
        # Obtener la actividad de la base de datos
        actividad_model = ActividadModel()
        actividad = actividad_model.obtener_actividad_por_id(actividad_id)
        
        if not actividad:
            logger.error(f"Actividad no encontrada: {actividad_id}")
            return jsonify({
                'mensaje': 'Actividad no encontrada',
                'error': f'No se encontró la actividad con ID {actividad_id}'
            }), 404
            
        logger.info(f"Actividad encontrada: {actividad.get('tema')}")
        logger.info(f"Número de asistentes: {len(actividad.get('asistentes', []))}")
        
        # Obtener los datos completos de los beneficiarios
        if 'asistentes' in actividad and actividad['asistentes']:
            beneficiario_model = BeneficiarioModel()
            for asistente in actividad['asistentes']:
                if 'beneficiario_id' in asistente and asistente['beneficiario_id']:
                    try:
                        logger.info(f"Buscando beneficiario con ID: {asistente['beneficiario_id']}")
                        beneficiario = beneficiario_model.obtener_beneficiario_por_id(asistente['beneficiario_id'])
                        if beneficiario:
                            logger.info(f"Beneficiario encontrado: {beneficiario.get('nombre_completo', 'Sin nombre')}")
                            asistente['beneficiario'] = beneficiario
                        else:
                            logger.warning(f"No se encontró el beneficiario con ID: {asistente['beneficiario_id']}")
                    except Exception as e:
                        logger.error(f"Error al obtener beneficiario {asistente['beneficiario_id']}: {str(e)}", exc_info=True)
                        continue
        
        if not actividad:
            logger.error(f"Actividad no encontrada: {actividad_id}")
            return jsonify({
                'success': False,
                'message': 'Actividad no encontrada'
            }), 404
            
        logger.info(f"Actividad encontrada: {actividad.get('tema', 'Sin nombre')}")
        
        # Crear un libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Definir todas las columnas posibles con sus etiquetas
        columnas_disponibles = [
            {'campo': 'numero', 'etiqueta': 'N°', 'visible_por_defecto': True},
            {'campo': 'fecha_registro', 'etiqueta': 'FECHA DE REGISTRO', 'visible_por_defecto': True},
            {'campo': 'nombre', 'etiqueta': 'NOMBRE', 'visible_por_defecto': True},
            {'campo': 'tipo_documento', 'etiqueta': 'TIPO DOCUMENTO', 'visible_por_defecto': True},
            {'campo': 'identificacion', 'etiqueta': 'IDENTIFICACIÓN', 'visible_por_defecto': True},
            {'campo': 'genero', 'etiqueta': 'GÉNERO', 'visible_por_defecto': True},
            {'campo': 'edad', 'etiqueta': 'EDAD', 'visible_por_defecto': True},
            {'campo': 'rango_edad', 'etiqueta': 'RANGO DE EDAD', 'visible_por_defecto': True},
            {'campo': 'nivel_educativo', 'etiqueta': 'NIVEL EDUCATIVO', 'visible_por_defecto': True},
            {'campo': 'comuna', 'etiqueta': 'COMUNA', 'visible_por_defecto': True},
            {'campo': 'barrio', 'etiqueta': 'BARRIO', 'visible_por_defecto': True},
            {'campo': 'telefono', 'etiqueta': 'TELÉFONO', 'visible_por_defecto': True},
            {'campo': 'correo', 'etiqueta': 'CORREO ELECTRÓNICO', 'visible_por_defecto': True},
            {'campo': 'estudia', 'etiqueta': 'ESTUDIA ACTUALMENTE', 'visible_por_defecto': False},
            {'campo': 'sabe_leer', 'etiqueta': 'SABE LEER', 'visible_por_defecto': False},
            {'campo': 'sabe_escribir', 'etiqueta': 'SABE ESCRIBIR', 'visible_por_defecto': False},
            {'campo': 'tipo_vivienda', 'etiqueta': 'TIPO DE VIVIENDA', 'visible_por_defecto': False},
            {'campo': 'situacion_laboral', 'etiqueta': 'SITUACIÓN LABORAL', 'visible_por_defecto': False},
            {'campo': 'grupo_etnico', 'etiqueta': 'GRUPO ÉTNICO', 'visible_por_defecto': False},
            {'campo': 'ayuda_humanitaria', 'etiqueta': 'AYUDA HUMANITARIA', 'visible_por_defecto': False},
            {'campo': 'tipo_ayuda_humanitaria', 'etiqueta': 'TIPO DE AYUDA HUMANITARIA', 'visible_por_defecto': False},
            {'campo': 'discapacidad', 'etiqueta': 'DISCAPACIDAD', 'visible_por_defecto': False},
            {'campo': 'tipo_discapacidad', 'etiqueta': 'TIPO DE DISCAPACIDAD', 'visible_por_defecto': False},
            {'campo': 'nombre_cuidadora', 'etiqueta': 'NOMBRE DEL CUIDADOR/A', 'visible_por_defecto': False},
            {'campo': 'labora_actualmente', 'etiqueta': 'LABORA ACTUALMENTE', 'visible_por_defecto': False},
            {'campo': 'victima_conflicto', 'etiqueta': 'VÍCTIMA DE CONFLICTO', 'visible_por_defecto': False},
            {'campo': 'qr_registro', 'etiqueta': 'CÓDIGO QR', 'visible_por_defecto': False}
        ]
        
        # Filtrar columnas según lo solicitado
        if columnas:
            # Filtrar solo las columnas que existen en columnas_disponibles
            columnas_seleccionadas = [col for col in columnas 
                                    if any(c['campo'] == col for c in columnas_disponibles)]
            
            # Si no hay columnas válidas, usar las visibles por defecto
            if not columnas_seleccionadas:
                columnas_seleccionadas = [c['campo'] for c in columnas_disponibles if c['visible_por_defecto']]
        else:
            # Si no se especifican columnas, usar las visibles por defecto
            columnas_seleccionadas = [c['campo'] for c in columnas_disponibles if c['visible_por_defecto']]
        
        # Obtener los encabezados para las columnas seleccionadas
        headers = []
        for col in columnas_disponibles:
            if col['campo'] in columnas_seleccionadas:
                headers.append(col['etiqueta'])
        
        # Agregar encabezados (comenzando en la fila 5)
        start_row = 5
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Ajustar ancho de columna según el contenido
            column_letter = get_column_letter(col_num)
            max_length = len(header)
            adjusted_width = (max_length + 2) * 1.2  # Ajuste del ancho
            
            # Asegurar un ancho mínimo y máximo
            if adjusted_width < 10:
                adjusted_width = 10
            elif adjusted_width > 40:  # Ancho máximo
                adjusted_width = 40
                ws.column_dimensions[column_letter].width = adjusted_width
                # Ajustar el alto de la fila para caber el texto
                ws.row_dimensions[start_row].height = 30
                ws.cell(row=start_row, column=col_num).alignment = Alignment(wrap_text=True, vertical='center')
            else:
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Mapeo de campos a sus valores correspondientes
        def obtener_valor_campo(asistente, campo, indice):
            beneficiario = asistente.get('beneficiario', {})
            
            # Obtener fecha de registro del asistente o del beneficiario
            fecha_registro = asistente.get('fecha_asistencia') or asistente.get('fecha_registro')
            if not fecha_registro:
                fecha_registro = beneficiario.get('fecha_registro', '')
            
            # Formatear la fecha si existe
            if fecha_registro:
                try:
                    # Si es un string, intentar parsearlo
                    if isinstance(fecha_registro, str):
                        # Intentar diferentes formatos de fecha
                        formatos_fecha = [
                            '%Y-%m-%dT%H:%M:%S.%fZ',  # Formato con milisegundos y Z
                            '%Y-%m-%dT%H:%M:%S',       # Formato sin milisegundos
                            '%Y-%m-%d %H:%M:%S',        # Formato con espacio
                            '%Y-%m-%d'                  # Solo fecha
                        ]
                        
                        fecha_obj = None
                        for formato in formatos_fecha:
                            try:
                                fecha_obj = datetime.strptime(fecha_registro.split('.')[0], formato)
                                break
                            except (ValueError, AttributeError):
                                continue
                        
                        if fecha_obj:
                            fecha_registro = fecha_obj.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            logger.warning(f"No se pudo formatear la fecha: {fecha_registro}")
                            fecha_registro = str(fecha_registro).split('.')[0]  # Eliminar milisegundos si existen
                    # Si ya es un objeto datetime
                    elif hasattr(fecha_registro, 'strftime'):
                        fecha_registro = fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
                except Exception as e:
                    logger.warning(f"Error al formatear fecha {fecha_registro}: {str(e)}")
                    fecha_registro = str(fecha_registro).split('.')[0]  # Eliminar milisegundos si existen
            
            # Obtener enlace QR si existe
            enlace_qr = ''
            try:
                # Buscar en huella_dactilar
                if beneficiario.get('huella_dactilar'):
                    if isinstance(beneficiario['huella_dactilar'], dict):
                        if beneficiario['huella_dactilar'].get('datos_biometricos', {}).get('enlace_qr'):
                            enlace_qr = beneficiario['huella_dactilar']['datos_biometricos']['enlace_qr']
                        elif beneficiario['huella_dactilar'].get('enlace_qr'):
                            enlace_qr = beneficiario['huella_dactilar']['enlace_qr']
                # Buscar en el nivel superior del beneficiario
                if not enlace_qr and 'qr_registro' in beneficiario:
                    enlace_qr = beneficiario['qr_registro']
            except Exception as e:
                logger.warning(f"Error al obtener enlace QR: {str(e)}")
            
            # Obtener rango de edad
            rango_edad = beneficiario.get('rango_edad', '')
            if not rango_edad and 'edad' in beneficiario:
                try:
                    edad = int(float(beneficiario['edad']))  # Manejar tanto strings como números
                    if edad < 6: rango_edad = "0-5"
                    elif edad < 12: rango_edad = "6-11"
                    elif edad < 18: rango_edad = "12-17"
                    elif edad < 26: rango_edad = "18-25"
                    elif edad < 36: rango_edad = "26-35"
                    elif edad < 46: rango_edad = "36-45"
                    elif edad < 56: rango_edad = "46-55"
                    elif edad < 66: rango_edad = "56-65"
                    else: rango_edad = "66+"
                    
                    # Actualizar el rango de edad en el beneficiario para futuras referencias
                    if 'beneficiario' in asistente:
                        asistente['beneficiario']['rango_edad'] = rango_edad
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error al calcular rango de edad: {str(e)}")
                    rango_edad = ''
            
            # Obtener valores según el campo solicitado
            mapeo_campos = {
                'numero': indice + 1,
                'fecha_registro': fecha_registro if fecha_registro else asistente.get('fecha_registro', ''),
                'nombre': beneficiario.get('nombre_completo', ''),
                'tipo_documento': beneficiario.get('tipo_documento', ''),
                'identificacion': str(beneficiario.get('numero_documento', '')),
                'genero': beneficiario.get('genero', ''),
                'edad': beneficiario.get('edad', ''),  # Mantener el campo edad original
                'rango_edad': rango_edad,  # Usar el rango calculado
                'comuna': beneficiario.get('comuna', ''),
                'barrio': beneficiario.get('barrio', ''),
                'telefono': beneficiario.get('numero_celular', ''),
                'correo': beneficiario.get('correo_electronico', ''),
                'estudia': 'Sí' if beneficiario.get('estudia_actualmente', False) else 'No',
                'nivel_educativo': beneficiario.get('nivel_educativo', ''),
                'sabe_leer': 'Sí' if beneficiario.get('sabe_leer', False) else 'No',
                'sabe_escribir': 'Sí' if beneficiario.get('sabe_escribir', False) else 'No',
                'tipo_vivienda': beneficiario.get('tipo_vivienda', ''),
                'situacion_laboral': beneficiario.get('situacion_laboral', ''),
                'grupo_etnico': beneficiario.get('etnia', ''),
                'ayuda_humanitaria': 'Sí' if beneficiario.get('ayuda_humanitaria', False) else 'No',
                'tipo_ayuda_humanitaria': beneficiario.get('descripcion_ayuda_humanitaria', ''),
                'discapacidad': 'Sí' if beneficiario.get('tiene_discapacidad', False) else 'No',
                'tipo_discapacidad': beneficiario.get('tipo_discapacidad', ''),
                'nombre_cuidadora': beneficiario.get('nombre_cuidadora', ''),
                'labora_actualmente': 'Sí' if beneficiario.get('labora_actualmente', False) else 'No',
                'victima_conflicto': 'Sí' if beneficiario.get('victima_conflicto', False) else 'No',
                'qr_registro': enlace_qr,
                'nivel_educativo': beneficiario.get('nivel_educativo', ''),
                'tipo_ayuda_humanitaria': beneficiario.get('tipo_ayuda_humanitaria', beneficiario.get('descripcion_ayuda_humanitaria', ''))
            }
            
            return mapeo_campos.get(campo, '')
        
        # Llenar datos (comenzando después de los encabezados)
        row_num = start_row + 1  # Comenzar después de los encabezados (fila 6)
        for i, asistente in enumerate(actividad.get('asistentes', [])):
            try:
                # Obtener valores para cada columna seleccionada
                row_data = []
                for col in columnas_disponibles:
                    if col['campo'] in columnas_seleccionadas:
                        valor = obtener_valor_campo(asistente, col['campo'], i)
                        row_data.append(valor)
                
                # Agregar fila a la hoja
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                row_num += 1
                
            except Exception as e:
                logger.error(f"Error al procesar asistente {i}: {str(e)}", exc_info=True)
                continue
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 50)  # Máximo 50 caracteres
        
        # Crear un archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            excel_path = tmp_file.name
        
        # Guardar el libro de Excel
        wb.save(excel_path)
        logger.info(f"Archivo Excel generado: {excel_path}")
        
        # Enviar el archivo
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=f"asistencia-{actividad.get('tema', 'actividad')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error al exportar actividad {actividad_id}", exc_info=True)
        logger.error(f"Tipo de error: {type(e).__name__}")
        logger.error(f"Mensaje de error: {str(e)}")
        
        # Si hay un error, asegurarse de eliminar el archivo temporal
        if 'excel_path' in locals() and os.path.exists(excel_path):
            try:
                os.unlink(excel_path)
                logger.info(f"Archivo temporal eliminado: {excel_path}")
            except Exception as e2:
                logger.error(f"Error al eliminar archivo temporal {excel_path}: {str(e2)}")
        
        return jsonify({
            'success': False,
            'message': f'Error al exportar la actividad: {str(e)}',
            'error_type': type(e).__name__,
            'error_details': str(e)
        }), 500

def subir_logo():
    try:
        from app.routes.auth import token_required
        
        if 'logo' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No se ha subido ningún archivo'
            }), 400
            
        file = request.files['logo']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': 'No se ha seleccionado ningún archivo'
            }), 400
            
        # Validar extensión
        allowed_extensions = {'png', 'jpg', 'jpeg', 'svg'}
        filename = secure_filename(file.filename)
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return jsonify({
                'success': False,
                'message': 'Extensión de archivo no permitida'
            }), 400
            
        # Validar tamaño (máx 2MB)
        if file.content_length > 2 * 1024 * 1024:
            return jsonify({
                'success': False,
                'message': 'El archivo es demasiado grande (máx 2MB)'
            }), 400
            
        # Crear directorio de uploads si no existe
        upload_folder = configurar_subida_logo()
        
        # Generar nombre único para el archivo
        unique_filename = f"{ObjectId()}_{filename}"
        filepath = os.path.join(upload_folder, unique_filename)
        
        # Guardar el archivo
        file.save(filepath)
        
        # Devolver la ruta relativa con formato URL
        rel_path = os.path.join('uploads', 'logos', unique_filename).replace('\\', '/')
        
        return jsonify({
            'success': True,
            'url': f'/{rel_path}',
            'message': 'Logo subido correctamente'
        })
        
    except Exception as e:
        logger.error(f"Error al subir logo: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Error al subir el logo',
            'error': str(e)
        }), 500

def generar_pdf(actividad, output_path):
    """Genera un PDF con la información de la actividad y sus asistentes"""
    try:
        logger.info("Iniciando generación de PDF...")
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from datetime import datetime
        
        logger.info("Creando documento PDF...")
        doc = canvas.Canvas(output_path, pagesize=letter)
        width, height = letter
        
        # Configuración de estilos
        styles = getSampleStyleSheet()
        
        # Logo de la aplicación
        try:
            frontend_path = os.path.abspath(os.path.join(current_app.root_path, '..', '..', 'frontend', 'src'))
            logo_path = os.path.join(frontend_path, 'fondo', 'logo.png')
            
            logger.info(f"Buscando logo en: {logo_path}")
            
            if os.path.exists(logo_path):
                # Ajustar el tamaño manteniendo la proporción
                logo_width = 100
                logo_height = 50
                logger.info("Dibujando logo en el PDF...")
                doc.drawImage(logo_path, 50, height-70, width=logo_width, height=logo_height, mask='auto')
                logger.info("Logo dibujado correctamente")
            else:
                logger.warning(f"No se encontró el logo en la ruta: {logo_path}")
        except Exception as e:
            logger.error(f"Error al cargar el logo: {str(e)}", exc_info=True)
        
        # Títulos estáticos
        logger.info("Agregando títulos al PDF...")
        doc.setFont("Helvetica-Bold", 18)
        doc.drawCentredString(width/2, height-50, "FORMATO REGISTRO DE ASISTENCIA")
        doc.setFont("Helvetica-Bold", 14)
        doc.drawCentredString(width/2, height-75, "ALCALDÍA MUNICIPAL DE QUIBDÓ")
        doc.setFont("Helvetica-Bold", 12)
        doc.drawCentredString(width/2, height-95, "LISTADO DE BENEFICIARIOS")
        
        # Fecha de generación
        doc.setFont("Helvetica", 10)
        doc.drawRightString(width-50, height-60, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        # Línea separadora
        doc.line(50, height-80, width-50, height-80)
        
        # Información de la actividad
        logger.info("Agregando información de la actividad...")
        doc.setFont("Helvetica-Bold", 12)
        doc.drawString(50, height-110, f"Actividad: {actividad.get('tema', 'Sin tema especificado')}")
        
        # Detalles de la actividad
        doc.setFont("Helvetica", 10)
        y_position = height-130
        detalles = [
            ("Lugar:", actividad.get('lugar', 'No especificado')),
            ("Fecha:", actividad.get('fecha', 'No especificada')),
            ("Hora Inicio:", actividad.get('hora_inicio', 'No especificada')),
            ("Hora Fin:", actividad.get('hora_fin', 'No especificada')),
            ("Línea de Trabajo:", actividad.get('linea_trabajo', {}).get('nombre', 'No especificada')),
            ("Responsable:", actividad.get('creado_por_nombre', 'No especificado'))
        ]
        
        for etiqueta, valor in detalles:
            doc.drawString(50, y_position, f"{etiqueta} {valor}")
            y_position -= 15
        
        # Espacio antes de la tabla
        y_position -= 30
        
        # Tabla de beneficiarios
        logger.info("Preparando tabla de beneficiarios...")
        
        # Encabezados de la tabla según los campos solicitados
        encabezados = [
            'N°', 'FECHA DE REGISTRO', 'NOMBRE', 'TIPO DOCUMENTO', 
            'IDENTIFICACIÓN', 'GÉNERO', 'RANGO DE EDAD', 'COMUNA', 'BARRIO',
            'CORREO ELECTRONICO', 'NUMERO CELULAR', 'ESTUDIA ACTUALMENTE',
            'NIVEL EDUCATIVO', 'SABE LEER', 'SABE ESCRIBIR', 'TIPO DE VIVIENDA',
            'SITUACION LABORAL', 'GRUPO ETNICO', 'AYUDA HUMANITARIA',
            'DISCAPACIDAD', 'TIPO DE DISCAPACIDAD', 'NOMBRE DE LA CUIDADORA',
            'LABORA ACTUALMENTE', 'VICTIMA DE CONFLICTO', 'QR DE REGISTRO'
        ]
        
        # Anchos de columna ajustados a los nuevos campos
        col_widths = [
            10, 30, 80, 30,  # N°, Fecha Registro, Nombre, Tipo Documento
            40, 30, 30, 40, 40,  # Identificación, Género, Rango Edad, Comuna, Barrio
            60, 40, 30,         # Correo, Celular, Estudia Actualmente
            40, 20, 20, 40,     # Nivel Educativo, Sabe Leer, Sabe Escribir, Tipo Vivienda
            40, 40, 30,         # Situación Laboral, Grupo Étnico, Ayuda Humanitaria
            20, 40, 50,         # Discapacidad, Tipo Discapacidad, Nombre Cuidadora
            30, 30, 30          # Labora Actualmente, Víctima Conflicto, QR
        ]
        
        # Crear datos de la tabla
        data = [encabezados]
        
        if 'asistentes' in actividad and actividad['asistentes']:
            logger.info(f"Procesando {len(actividad['asistentes'])} asistentes...")
            for i, asistente in enumerate(actividad['asistentes'], 1):
                try:
                    beneficiario = asistente.get('beneficiario', {})
                    
                    # Obtener datos del beneficiario según los campos solicitados
                    fecha_registro = asistente.get('fecha_asistencia', '')
                    if fecha_registro and isinstance(fecha_registro, str):
                        try:
                            # Formatear la fecha si es necesario
                            fecha_obj = datetime.strptime(fecha_registro, '%Y-%m-%dT%H:%M:%S.%fZ')
                            fecha_registro = fecha_obj.strftime('%d/%m/%Y')
                        except (ValueError, TypeError):
                            fecha_registro = fecha_registro.split('T')[0]  # Tomar solo la parte de la fecha
                    
                    # Obtener rango de edad basado en la edad
                    edad = int(beneficiario.get('edad', 0)) if str(beneficiario.get('edad', '')).isdigit() else 0
                    rango_edad = (
                        '0-5' if edad <= 5 else
                        '6-12' if edad <= 12 else
                        '13-17' if edad <= 17 else
                        '18-28' if edad <= 28 else
                        '29-59' if edad <= 59 else
                        '60+'
                    )
                    
                    fila = [
                        str(i),  # N°
                        fecha_registro,  # FECHA DE REGISTRO
                        beneficiario.get('nombre_completo', 'N/A'),  # NOMBRE
                        beneficiario.get('tipo_documento', ''),  # TIPO DOCUMENTO
                        str(beneficiario.get('numero_documento', '')),  # IDENTIFICACIÓN
                        beneficiario.get('genero', ''),  # GÉNERO
                        rango_edad,  # RANGO DE EDAD
                        beneficiario.get('comuna', ''),  # COMUNA
                        beneficiario.get('barrio', ''),  # BARRIO
                        beneficiario.get('correo_electronico', ''),  # CORREO ELECTRONICO
                        beneficiario.get('numero_celular', ''),  # NUMERO CELULAR
                        'Sí' if beneficiario.get('estudia_actualmente', False) else 'No',  # ESTUDIA ACTUALMENTE
                        beneficiario.get('nivel_educativo', ''),  # NIVEL EDUCATIVO
                        'Sí' if beneficiario.get('sabe_leer', False) else 'No',  # SABE LEER
                        'Sí' if beneficiario.get('sabe_escribir', False) else 'No',  # SABE ESCRIBIR
                        beneficiario.get('tipo_vivienda', ''),  # TIPO DE VIVIENDA
                        beneficiario.get('situacion_laboral', ''),  # SITUACION LABORAL
                        beneficiario.get('grupo_etnico', ''),  # GRUPO ETNICO
                        'Sí' if beneficiario.get('ayuda_humanitaria', False) else 'No',  # AYUDA HUMANITARIA
                        'Sí' if beneficiario.get('tiene_discapacidad', False) else 'No',  # DISCAPACIDAD
                        beneficiario.get('tipo_discapacidad', ''),  # TIPO DE DISCAPACIDAD
                        beneficiario.get('nombre_cuidador', ''),  # NOMBRE DE LA CUIDADORA
                        'Sí' if beneficiario.get('labora_actualmente', False) else 'No',  # LABORA ACTUALMENTE
                        'Sí' if beneficiario.get('victima_conflicto', False) else 'No',  # VICTIMA DE CONFLICTO
                        ''  # QR DE REGISTRO
                    ]
                    
                    # Agregar imagen del código QR si existe
                    if 'huella_dactilar' in beneficiario and 'enlace_qr' in beneficiario['huella_dactilar']:
                        qr_path = beneficiario['huella_dactilar']['enlace_qr']
                        if os.path.exists(qr_path):
                            # Guardar la posición para dibujar el QR después
                            fila[-1] = f"QR_{i}"
                            qr_imagenes[f"QR_{i}"] = qr_path
                    data.append(fila)
                except Exception as e:
                    logger.error(f"Error al procesar beneficiario {i}: {str(e)}")
                    continue
        else:
            logger.info("No hay beneficiarios para mostrar")
        
        # Crear tabla
        logger.info("Creando tabla...")
        try:
            # Ajustar el ancho de la tabla al ancho de la página
            table_width = sum(col_widths)
            x_position = (width - table_width) / 2
            
            # Crear la tabla con los anchos de columna personalizados
            table = Table(data, colWidths=col_widths)
            
            # Estilos de la tabla
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),  # Color de fondo del encabezado
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Color del texto del encabezado
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alineación centrada
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fuente en negrita para el encabezado
                ('FONTSIZE', (0, 0), (-1, 0), 8),  # Tamaño de fuente del encabezado
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),  # Espaciado inferior del encabezado
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Fondo blanco para el contenido
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Color del texto del contenido
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Fuente normal para el contenido
                ('FONTSIZE', (0, 1), (-1, -1), 8),  # Tamaño de fuente del contenido
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alineación vertical al centro
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),  # Líneas de la tabla
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),  # Borde de la tabla
            ]))
            
            # Ajustar el alto de las filas
            for i in range(1, len(data)):
                table.setStyle(TableStyle([
                    ('FONTSIZE', (0, i), (-1, i), 8),  # Tamaño de fuente para todas las celdas de la fila
                    ('LEFTPADDING', (0, i), (-1, i), 3),  # Espaciado izquierdo
                    ('RIGHTPADDING', (0, i), (-1, i), 3),  # Espaciado derecho
                    ('TOPPADDING', (0, i), (-1, i), 3),  # Espaciado superior
                    ('BOTTOMPADDING', (0, i), (-1, i), 3),  # Espaciado inferior
                ]))
            
            # Dibujar tabla
            logger.info("Dibujando tabla en el PDF...")
            table.wrapOn(doc, width, height)
            table.drawOn(doc, x_position, y_position - (len(data) * 15) - 30)  # Ajustar posición Y
            
            # Dibujar códigos QR después de la tabla
            current_y = y_position - (len(data) * 15) - 30
            for i, row in enumerate(data[1:], 1):  # Saltar encabezado
                if row[-1].startswith('QR_'):
                    qr_path = qr_imagenes.get(row[-1])
                    if qr_path and os.path.exists(qr_path):
                        try:
                            # Calcular posición X para el código QR (última columna)
                            qr_x = x_position + sum(col_widths[:-1]) + 5
                            # Calcular posición Y para la fila actual
                            qr_y = current_y - (i * 15) - 5
                            # Dibujar código QR (15x15 mm)
                            doc.drawImage(qr_path, qr_x, qr_y, width=15, height=15)
                        except Exception as e:
                            logger.error(f"Error al dibujar código QR: {str(e)}")
            
            # Pie de página
            logger.info("Agregando pie de página...")
            doc.setFont("Helvetica", 10)
            doc.drawCentredString(width/2, 50, "________________________________________")
            doc.drawCentredString(width/2, 35, "Firma del Responsable")
            
            # Información de la organización
            doc.setFont("Helvetica", 8)
            doc.drawCentredString(width/2, 20, "Red de Inclusión Social - Todos los derechos reservados")
            
            # Guardar PDF
            logger.info("Guardando PDF...")
            doc.save()
            logger.info("PDF generado exitosamente")
            
        except Exception as e:
            logger.error(f"Error al crear la tabla: {str(e)}", exc_info=True)
            raise Exception(f"Error al generar la tabla del PDF: {str(e)}")
            
    except Exception as e:
        logger.error(f"Error crítico al generar el PDF: {str(e)}", exc_info=True)
        # Asegurarse de que el documento se cierre correctamente en caso de error
        if 'doc' in locals():
            try:
                doc.save()
            except:
                pass
        raise  # Relanzar la excepción para que sea manejada por el llamante