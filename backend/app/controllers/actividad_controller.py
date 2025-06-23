# backend/app/controllers/actividad_controller.py
from flask import request, current_app, jsonify, send_file, g
from bson import ObjectId
from werkzeug.utils import secure_filename
from datetime import datetime, timezone
import os
import logging
import traceback
import tempfile  # Para manejo de archivos temporales
import requests  # Para descargar imágenes de códigos QR existentes
import time     # Para manejo de timestamps
from io import BytesIO  # Para manejar datos binarios en memoria
from bson.errors import InvalidId
from app.models.actividad import ActividadModel, actividad_schema, ActividadSchema
from PIL import Image
import base64

# Configurar el logger
logger = logging.getLogger(__name__)

def configurar_subida_logo():
    upload_folder = os.path.join(current_app.root_path, '..', '..', 'uploads', 'logos')
    os.makedirs(upload_folder, exist_ok=True)
    return upload_folder

def crear_actividad():
    try:
        from app.models.actividad import ActividadModel, actividad_schema, ActividadSchema
        
        # Obtener datos del formulario
        datos = request.get_json()
        if not datos:
            logger.error("No se recibieron datos en la solicitud")
            return jsonify({
                'success': False,
                'message': 'No se recibieron datos en la solicitud'
            }), 400
            
        logger.info(f"Datos recibidos en el backend: {datos}")
        logger.info(f"Headers: {dict(request.headers)}")
            
        logger.info(f"Datos recibidos para crear actividad: {datos}")
        
        # Validar que los datos sean un diccionario
        if not isinstance(datos, dict):
            logger.error(f"Los datos deben ser un diccionario, se recibió: {type(datos)}")
            return jsonify({
                'success': False,
                'message': 'Los datos deben ser un diccionario',
                'tipo_recibido': str(type(datos))
            }), 400
        
        # Validar que todos los campos requeridos estén presentes
        required_fields = [
            'tema', 'objetivo', 'lugar', 'dependencia', 'fecha',
            'hora_inicio', 'hora_fin', 'linea_trabajo_id', 'tipo'
        ]
        
        # Confiar en el valor de 'tipo' que viene del frontend
        logger.info(f"Tipo recibido del frontend: {datos.get('tipo')}")
        logger.info(f"Datos completos recibidos: {datos}")
        
        logger.info(f"Datos a guardar - Tipo: {datos.get('tipo')}")
        
        # Si no hay creado_por, usar un valor por defecto
        if 'creado_por' not in datos:
            datos['creado_por'] = 'sistema'  # O cualquier otro valor por defecto
            logger.info("No se proporcionó 'creado_por', usando valor por defecto")
        
        # Asegurarse de que funcionario_id esté presente, si no, usar creado_por
        if 'funcionario_id' not in datos:
            datos['funcionario_id'] = datos['creado_por']
            logger.info(f"Usando creado_por como funcionario_id: {datos['funcionario_id']}")
        
        # Verificar campos faltantes
        missing_fields = [field for field in required_fields if field not in datos]
        if missing_fields:
            logger.error(f"Faltan campos requeridos: {missing_fields}")
            return jsonify({
                'success': False,
                'message': 'Faltan campos requeridos',
                'missing_fields': missing_fields,
                'datos_recibidos': {k: v for k, v in datos.items() if k in required_fields}
            }), 422
            
        # Validar formato de fecha
        try:
            fecha_str = datos['fecha']
            if isinstance(fecha_str, str):
                # Si es string, convertirlo a datetime
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
                # Asegurarse de que la fecha esté en UTC
                if hasattr(fecha_obj, 'tzinfo') and fecha_obj.tzinfo is not None:
                    fecha_obj = fecha_obj.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                datos['fecha'] = fecha_obj
            elif isinstance(fecha_str, datetime):
                # Si ya es datetime, asegurarse de que esté en UTC
                if hasattr(fecha_str, 'tzinfo') and fecha_str.tzinfo is not None:
                    fecha_obj = fecha_str.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                else:
                    fecha_obj = fecha_str
                datos['fecha'] = fecha_obj
            else:
                raise ValueError(f"Formato de fecha no soportado: {type(fecha_str)}")
                
            logger.info(f"Fecha convertida: {datos['fecha']} ({type(datos['fecha']).__name__})")
        except Exception as e:
            logger.error(f"Error al procesar la fecha: {str(e)}")
            return jsonify({
                'success': False,
                'message': 'Error en el formato de fecha',
                'error': str(e)
            }), 400
            
        # Validar formato de horas
        try:
            datetime.strptime(datos['hora_inicio'], '%H:%M')
            datetime.strptime(datos['hora_fin'], '%H:%M')
        except (ValueError, TypeError) as e:
            logger.error(f"Error en formato de hora: {str(e)}")
            return jsonify({
                'success': False,
                'message': 'Formato de hora inválido. Use HH:MM',
                'error': str(e)
            }), 422
        
        # Usar el usuario del sistema como creador si no se proporciona
        if 'creado_por' not in datos:
            datos['creado_por'] = 'sistema'
                
        # Usar creado_por como funcionario_id si no se proporciona
        if 'funcionario_id' not in datos:
            datos['funcionario_id'] = datos['creado_por']
        
        # Validar los datos con el esquema
        try:
            # Asegurarse de que la fecha sea un string ISO 8601
            if 'fecha' in datos and isinstance(datos['fecha'], datetime):
                datos['fecha'] = datos['fecha'].isoformat()
                
            # Asegurarse de que los IDs sean strings
            for field in ['linea_trabajo_id', 'funcionario_id', 'creado_por']:
                if field in datos and datos[field]:
                    if isinstance(datos[field], ObjectId):
                        datos[field] = str(datos[field])
                    elif not isinstance(datos[field], str):
                        datos[field] = str(datos[field])
            
            logger.info(f"Datos antes de validar con esquema: {datos}")
            
            # Validar manualmente el esquema
            schema = ActividadSchema()
            errors = schema.validate(datos)
            
            if errors:
                logger.error(f"Errores de validación: {errors}")
                return jsonify({
                    'success': False,
                    'message': 'Error de validación de datos',
                    'errors': errors,
                    'datos_recibidos': datos
                }), 422
                
            # Si pasa la validación, cargar los datos
            result = schema.load(datos)
            logger.info(f"Datos validados correctamente: {result}")
        except Exception as e:
            error_details = {
                'error': str(e),
                'error_type': type(e).__name__,
                'validation_errors': getattr(e, 'messages', 'No hay mensajes de validación'),
                'datos_recibidos': datos
            }
            logger.error(f"Error de validación: {error_details}")
            return jsonify({
                'success': False,
                'message': 'Error de validación de datos',
                **error_details
            }), 422
        
        # Crear la actividad
        actividad_model = ActividadModel()
        actividad_id = actividad_model.crear_actividad(datos)
        
        if not actividad_id:
            raise Exception("No se pudo crear la actividad")
            
        return jsonify({
            'success': True,
            'message': 'Actividad creada exitosamente',
            'actividad_id': str(actividad_id)
        }), 201
        
    except Exception as e:
        logger.error(f"Error al crear actividad: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'Error al crear la actividad',
            'error': str(e),
            'error_type': type(e).__name__
        }), 500

def obtener_actividades():
    try:
        from app.models.actividad import ActividadModel
        
        # Obtener filtros de la URL
        filtros = request.args.to_dict()
        
        # Inicializar el modelo
        actividad_model = ActividadModel()
        
        # Obtener actividades con los filtros proporcionados
        actividades = actividad_model.obtener_actividades(filtros)
        
        return jsonify({
            'success': True,
            'data': actividades
        })
        
    except Exception as e:
        logger.error(f"Error al obtener actividades: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Error al obtener las actividades',
            'error': str(e)
        }), 500

def obtener_actividad(actividad_id):
    """
    Obtiene los detalles de una actividad específica por su ID
    """
    from flask import jsonify
    from bson import ObjectId
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Obteniendo detalles de la actividad {actividad_id}")
        
        # Validar el ID de la actividad
        if not ObjectId.is_valid(actividad_id):
            logger.error(f"ID de actividad inválido: {actividad_id}")
            return jsonify({
                'success': False,
                'message': 'ID de actividad inválido'
            }), 400
        
        # Importar modelos dentro del bloque try para evitar importaciones circulares
        from app.models.actividad import ActividadModel
        from app.models.beneficiario import BeneficiarioModel
        
        # Obtener la actividad
        actividad_model = ActividadModel()
        actividad = actividad_model.obtener_actividad_por_id(actividad_id)
        
        if not actividad:
            logger.warning(f"No se encontró la actividad con ID: {actividad_id}")
            return jsonify({
                'success': False,
                'message': 'Actividad no encontrada'
            }), 404
        
        # Obtener información detallada de los beneficiarios si hay asistentes
        if 'asistentes' in actividad and actividad['asistentes']:
            logger.info(f"Obteniendo información de {len(actividad['asistentes'])} asistentes")
            aistente_model = BeneficiarioModel()
            
            for asistente in actividad['asistentes']:
                try:
                    beneficiario_id = asistente.get('beneficiario_id')
                    if not beneficiario_id:
                        logger.warning("Asistente sin beneficiario_id")
                        continue
                        
                    beneficiario = aistente_model.obtener_beneficiario_por_id(beneficiario_id)
                    if beneficiario:
                        # Agregar información del beneficiario al asistente
                        asistente['beneficiario'] = {
                            'nombre_completo': beneficiario.get('nombre_completo', ''),
                            'tipo_documento': beneficiario.get('tipo_documento', ''),
                            'numero_documento': beneficiario.get('numero_documento', '')
                        }
                except Exception as e:
                    logger.error(f"Error al obtener información del beneficiario: {str(e)}")
                    continue
        
        # Retornar la actividad con la información de los asistentes
        return jsonify({
            'success': True,
            'data': actividad
        })
        
    except Exception as e:
        logger.error(f"Error al obtener la actividad: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'Error al obtener la actividad',
            'error': str(e)
        }), 500

def actualizar_actividad(actividad_id):
    """
    Actualiza una actividad existente
    """
    logger.info(f"[ACTUALIZAR_ACTIVIDAD] Iniciando actualización de actividad {actividad_id}")
    
    try:
        # Verificar si hay datos JSON
        if not request.is_json:
            logger.error("[ACTUALIZAR_ACTIVIDAD] No se recibieron datos JSON en la solicitud")
            return jsonify({
                'success': False,
                'message': 'Se esperaba un cuerpo JSON en la solicitud',
                'error': 'NO_JSON_DATA'
            }), 400
            
        # Obtener datos JSON
        try:
            datos = request.get_json()
            logger.info(f"[ACTUALIZAR_ACTIVIDAD] Datos recibidos: {datos}")
            
            # Validar que los datos sean un diccionario
            if not isinstance(datos, dict):
                logger.error(f"[ACTUALIZAR_ACTIVIDAD] Los datos deben ser un objeto JSON. Tipo recibido: {type(datos)}")
                return jsonify({
                    'success': False,
                    'message': 'Los datos deben ser un objeto JSON',
                    'error': 'INVALID_JSON_FORMAT'
                }), 400
                
            # Validar que el tipo sea válido si está presente
            if 'tipo' in datos and datos['tipo'] not in ['actividad', 'reunion']:
                logger.warning(f"[ACTUALIZAR_ACTIVIDAD] Tipo de actividad no válido: {datos['tipo']}. No se actualizará el campo 'tipo'.")
                datos.pop('tipo', None)
                
        except Exception as e:
            error_msg = f"[ACTUALIZAR_ACTIVIDAD] Error al procesar los datos JSON: {str(e)}"
            logger.error(error_msg)
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Traceback: {traceback.format_exc()}")
            return jsonify({
                'success': False,
                'message': 'Error al procesar los datos de la solicitud',
                'error': 'INVALID_JSON',
                'details': str(e)
            }), 400
            
        if not datos:
            logger.error("[ACTUALIZAR_ACTIVIDAD] No se recibieron datos para actualizar")
            return jsonify({
                'success': False,
                'message': 'No se recibieron datos para actualizar',
                'error': 'EMPTY_DATA'
            }), 400
            
        # Validar que el ID sea un ObjectId válido
        try:
            if not ObjectId.is_valid(actividad_id):
                logger.error(f"[ACTUALIZAR_ACTIVIDAD] ID de actividad inválido: {actividad_id}")
                return jsonify({
                    'success': False,
                    'message': 'ID de actividad inválido',
                    'campo': 'id',
                    'error': 'INVALID_ACTIVITY_ID'
                }), 400
        except Exception as e:
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Error al validar el ID de actividad: {str(e)}")
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Traceback: {traceback.format_exc()}")
            return jsonify({
                'success': False,
                'message': 'Error al validar el ID de actividad',
                'error': 'VALIDATION_ERROR',
                'details': str(e)
            }), 500
    
        # Crear un nuevo diccionario con los campos permitidos
        datos_actualizados = {}
        
        # Lista de campos permitidos para actualización
        campos_permitidos = [
            'tema', 'objetivo', 'lugar', 'dependencia', 'fecha',
            'hora_inicio', 'hora_fin', 'linea_trabajo_id', 'estado',
            'asistentes', 'logo_url', 'actualizado_por', 'funcionario_id'
        ]
        
        # Copiar solo los campos permitidos que existen en los datos
        for campo in campos_permitidos:
            if campo in datos and datos[campo] is not None:  # Solo copiar si el campo existe y no es None
                datos_actualizados[campo] = datos[campo]
        
        # Si no hay campos para actualizar
        if not datos_actualizados:
            logger.warning("[ACTUALIZAR_ACTIVIDAD] No se proporcionaron campos válidos para actualizar")
            return jsonify({
                'success': True,
                'message': 'No se realizaron cambios (sin campos válidos para actualizar)',
                'modificados': 0
            })
        
        logger.info(f"[ACTUALIZAR_ACTIVIDAD] Datos a actualizar: {datos_actualizados}")
        
        # Validar campos requeridos solo si se están actualizando
        campos_requeridos = {
            'tema': 'El tema es requerido',
            'objetivo': 'El objetivo es requerido',
            'lugar': 'El lugar es requerido',
            'dependencia': 'La dependencia es requerida',
            'fecha': 'La fecha es requerida',
            'hora_inicio': 'La hora de inicio es requerida',
            'hora_fin': 'La hora de fin es requerida',
            'linea_trabajo_id': 'La línea de trabajo es requerida',
            'funcionario_id': 'El funcionario es requerido'
        }
        
        # Validar solo los campos que se están actualizando
        for campo, mensaje in campos_requeridos.items():
            if campo in datos_actualizados:  # Solo validar campos que se están actualizando
                if not datos_actualizados[campo]:
                    logger.error(f"[ACTUALIZAR_ACTIVIDAD] Campo requerido vacío: {campo}")
                    return jsonify({
                        'success': False,
                        'message': mensaje,
                        'campo': campo,
                        'error': 'EMPTY_REQUIRED_FIELD'
                    }), 400
        
        # Inicializar el modelo de actividad
        try:
            from app.models.actividad import ActividadModel
            actividad_model = ActividadModel()
            logger.info("[ACTUALIZAR_ACTIVIDAD] Modelo de actividad inicializado correctamente")
            
            # Verificar si la actividad existe
            logger.info(f"[ACTUALIZAR_ACTIVIDAD] Buscando actividad con ID: {actividad_id}")
            actividad_existente = actividad_model.obtener_actividad_por_id(actividad_id)
            
            if not actividad_existente:
                logger.error(f"[ACTUALIZAR_ACTIVIDAD] No se encontró la actividad con ID: {actividad_id}")
                return jsonify({
                    'success': False,
                    'message': 'No se encontró la actividad especificada',
                    'error': 'ACTIVITY_NOT_FOUND'
                }), 404
                
            # Agregar fecha de actualización
            datos_actualizados['fecha_actualizacion'] = datetime.utcnow()
            logger.info(f"[ACTUALIZAR_ACTIVIDAD] Fecha de actualización: {datos_actualizados['fecha_actualizacion']}")
            
            # Actualizar la actividad
            logger.info("[ACTUALIZAR_ACTIVIDAD] Actualizando actividad...")
            resultado = actividad_model.actualizar_actividad(actividad_id, datos_actualizados)
            
            if not resultado or not hasattr(resultado, 'modified_count') or resultado.modified_count == 0:
                logger.warning("[ACTUALIZAR_ACTIVIDAD] No se realizaron cambios en la actividad")
                return jsonify({
                    'success': True,
                    'message': 'No se realizaron cambios en la actividad',
                    'modificados': 0
                })
                
            logger.info(f"[ACTUALIZAR_ACTIVIDAD] Actividad {actividad_id} actualizada exitosamente")
            return jsonify({
                'success': True,
                'message': 'Actividad actualizada exitosamente',
                'modificados': resultado.modified_count
            })
            
        except Exception as e:
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Error al actualizar la actividad en la base de datos: {str(e)}")
            logger.error(f"[ACTUALIZAR_ACTIVIDAD] Traceback: {traceback.format_exc()}")
            return jsonify({
                'success': False,
                'message': 'Error al actualizar la actividad en la base de datos',
                'error': 'DATABASE_ERROR',
                'details': str(e)
            }), 500
            
    except Exception as e:
        logger.error(f"[ACTUALIZAR_ACTIVIDAD] Error inesperado al actualizar la actividad: {str(e)}")
        logger.error(f"[ACTUALIZAR_ACTIVIDAD] Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': 'Error inesperado al actualizar la actividad',
            'error': 'INTERNAL_SERVER_ERROR',
            'details': str(e)
        }), 500

def eliminar_actividad(actividad_id):
    try:
        from app.models.actividad import ActividadModel
        
        actividad_model = ActividadModel()
        eliminados = actividad_model.eliminar_actividad(actividad_id)
        
        if eliminados == 0:
            return jsonify({
                'success': False,
                'message': 'No se encontró la actividad'
            }), 404
            
        return jsonify({
            'success': True,
            'message': 'Actividad eliminada exitosamente',
            'eliminados': eliminados
        })
        
    except Exception as e:
        logger.error(f"Error al eliminar la actividad: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Error al eliminar la actividad',
            'error': str(e)
        }), 500

def registrar_asistencias(actividad_id):
    try:
        from app.models.actividad import ActividadModel
        from app.routes.auth import get_jwt_identity
        
        usuario_actual = get_jwt_identity()
        datos = request.get_json()
        
        if 'asistentes' not in datos:
            return jsonify({
                'success': False,
                'message': 'Se requiere la lista de asistentes'
            }), 400
            
        # Agregar información de actualización
        datos['actualizado_por'] = usuario_actual['id']
        
        actividad_model = ActividadModel()
        modificados = actividad_model.registrar_asistencias(actividad_id, datos['asistentes'])
        
        return jsonify({
            'success': True,
            'message': 'Asistencias registradas exitosamente',
            'modificados': modificados
        })
        
    except Exception as e:
        logger.error(f"Error al registrar asistencias: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Error al registrar asistencias',
            'error': str(e)
        }), 500
        

def exportar_actividad(actividad_id, columnas=None):
    try:
        from app.models.actividad import ActividadModel
        from app.models.beneficiario import BeneficiarioModel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        import tempfile
        import os
        import requests
        from io import BytesIO
        from bson import ObjectId
        from datetime import datetime
        from PIL import Image as PILImage
        from PIL import ImageDraw, ImageFont
        
        logger.info(f"Iniciando exportación de actividad a Excel: {actividad_id}")
        logger.info(f"Columnas seleccionadas: {columnas}")
        
        # Crear un libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"
        
        # Obtener la actividad de la base de datos
        actividad_model = ActividadModel()
        actividad = actividad_model.obtener_actividad_por_id(actividad_id)
        
        if not actividad:
            logger.error(f"Actividad no encontrada: {actividad_id}")
            return jsonify({
                'mensaje': 'Actividad no encontrada',
                'error': f'No se encontró la actividad con ID {actividad_id}'
            }), 404
        
        # Obtener los datos completos de los beneficiarios
        if 'asistentes' in actividad and actividad['asistentes']:
            beneficiario_model = BeneficiarioModel()
            for asistente in actividad['asistentes']:
                if 'beneficiario_id' in asistente and asistente['beneficiario_id']:
                    try:
                        logger.info(f"Buscando beneficiario con ID: {asistente['beneficiario_id']}")
                        beneficiario = beneficiario_model.obtener_beneficiario_por_id(asistente['beneficiario_id'])
                        if beneficiario:
                            asistente['beneficiario'] = beneficiario
                        else:
                            logger.warning(f"No se encontró el beneficiario con ID: {asistente['beneficiario_id']}")
                    except Exception as e:
                        logger.error(f"Error al obtener beneficiario {asistente['beneficiario_id']}: {str(e)}", exc_info=True)
                        continue
        
        if not actividad:
            logger.error(f"Actividad no encontrada: {actividad_id}")
            return jsonify({
                'success': False,
                'message': 'Actividad no encontrada'
            }), 404
            
        logger.info(f"Actividad encontrada: {actividad.get('tema', 'Sin nombre')}")
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Color de fondo para celdas de información (blanco)
        info_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Color de fondo para celdas de etiquetas (gris claro)
        label_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Agregar encabezado con logo e información de la actividad
        try:
            # Ruta al logo (usando ruta absoluta)
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            logo_path = os.path.join(base_dir, 'frontend', 'src', 'fondo', 'logo.png')
            logger.info(f"Ruta completa del logo: {logo_path}")
            logger.info(f"El archivo existe: {os.path.exists(logo_path)}")
            
            # Configuración de estilos
            font_size = 10
            font_name = 'Arial'
            
            # Configurar dimensiones de celdas
            ws.column_dimensions['A'].width = 15  # Columna A para el logo
            ws.column_dimensions['B'].width = 13  # Columna B para etiquetas
            ws.column_dimensions['C'].width = 20  # Columna C para valores
            ws.column_dimensions['D'].width = 15  # Columna D
            ws.column_dimensions['E'].width = 13  # Columna E
            ws.column_dimensions['F'].width = 15  # Columna F
            ws.column_dimensions['G'].width = 15  # Columna G
            ws.column_dimensions['H'].width = 13  # Columna H
            ws.column_dimensions['I'].width = 15  # Columna I
            ws.column_dimensions['J'].width = 15  # Columna J
            ws.column_dimensions['K'].width = 20  # Columna K
            
            # Altura de filas (se deja la altura por defecto)
            # No establecemos altura fija para que use la altura por defecto
            
            # Agregar logo (ocupa A1:C4)
            logo_size = 300  # Tamaño del logo más grande
            
            if os.path.exists(logo_path):
                try:
                    img = XLImage(logo_path)
                    # Ajustar tamaño manteniendo la relación de aspecto
                    img_ratio = img.height / img.width
                    # Redimensionar manteniendo la relación de aspecto
                    img.width = logo_size
                    img.height = int(logo_size * img_ratio)
                    
                    # No ajustamos la altura de las filas, se mantiene la altura por defecto
                    
                    # Agregar el logo
                    ws.add_image(img, 'A1')
                    logger.info("Logo agregado exitosamente")
                    
                    # Fusionar celdas para el logo (4 filas de alto x 3 columnas de ancho)
                    ws.merge_cells('A1:C4')
                    
                except Exception as e:
                    logger.error(f"Error al cargar la imagen: {str(e)}")
            else:
                logger.warning(f"No se encontró el archivo de logo en: {logo_path}")
            
            # Título principal (D1:K1)
            ws.merge_cells('D1:K1')
            title_cell = ws['D1']
            title_cell.value = 'FORMATO REGISTRO DE ASISTENCIA'
            title_cell.font = Font(bold=True, size=14, name='Arial')
            title_cell.fill = info_fill
            # title_cell.border = thin_border
            for row in ws['D1:K1']:
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Subtítulo (D2:K2)
            ws.merge_cells('D2:K2')
            subtitle_cell = ws['D2']
            subtitle_cell.value = 'ALCALDÍA MUNICIPAL DE QUIBDÓ'
            subtitle_cell.font = Font(bold=True, size=12, name='Arial')
            subtitle_cell.fill = info_fill
            for row in ws['D2:K2']:
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
            subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
                        

            # Fila 4: Dependencia
            ws['D3'].value = 'DEPENDENCIA:'
            ws['D3'].font = Font(bold=True, name='Arial', size=10)
            ws['D3'].fill = label_fill
            ws['D3'].border = thin_border
            ws.merge_cells('D3:E3')  # Fusionar D4:E4

            ws['F3'].value = actividad.get('dependencia', '').upper()
            ws['F3'].font = Font(name='Arial', size=10)
            ws['F3'].fill = info_fill
            ws['F3'].border = thin_border
            ws.merge_cells('F3:K3')  # Fusionar F4:I4

               # Fila 4: Dependencia
      
            # Combinar celdas sin texto
            ws.merge_cells('D4:K4')  # Fusionar D4:K4
            ws['D4'].fill = label_fill
            ws['D4'].border = thin_border

            # Fila 5: Tema
            ws['A5'].value = 'TEMA:'
            ws['A5'].font = Font(bold=True, name='Arial', size=10)
            ws['A5'].fill = label_fill
            ws['A5'].border = thin_border
            ws.merge_cells('A5:B5') 

            ws['C5'].value = actividad.get('tema', '').upper()
            ws['C5'].font = Font(name='Arial', size=10)
            ws['C5'].fill = info_fill
            ws['C5'].border = thin_border
            ws.merge_cells('C5:f5')  # Fusionar C5:F5

            # Fila 6: Fecha
            ws['A6'].value = 'FECHA:'
            ws['A6'].font = Font(bold=True, name='Arial', size=10)
            ws['A6'].fill = label_fill
            ws['A6'].border = thin_border
            ws.merge_cells('A6:B6')

          

            if 'fecha' in actividad:
                fecha_obj = datetime.strptime(actividad['fecha'], '%Y-%m-%dT%H:%M:%S')  # Si viene como string
                # O si ya es un objeto datetime: fecha_obj = reunion['fecha']
                fecha_formateada = fecha_obj.strftime('%d/%m/%Y')  # Formato DD/MM/YYYY
                ws['C6'].value = f"FECHA: {fecha_formateada}"
                ws['C6'].font = Font(name='Arial', size=10)
                ws['C6'].fill = info_fill
                ws['C6'].border = thin_border
            ws.merge_cells('C6') 

            # Fila 6: Lugar
            ws['D6'].value = 'LUGAR:'
            ws['D6'].font = Font(bold=True, name='Arial', size=10)
            ws['D6'].fill = label_fill
            ws['D6'].border = thin_border
            ws.merge_cells('D6') 

            ws['E6'].value = actividad.get('lugar', '').upper()
            ws['E6'].font = Font(name='Arial', size=10)
            ws['E6'].fill = info_fill
            ws['E6'].border = thin_border
            ws.merge_cells('E6:F6')  # Fusionar E6:F6

            # Fila 7: Hora de inicio
            ws['A7'].value = 'HORA INICIO:'
            ws['A7'].font = Font(bold=True, name='Arial', size=10)
            ws['A7'].fill = label_fill
            ws['A7'].border = thin_border
            ws.merge_cells('A7:B7') 

            ws['C7'].value = actividad.get('hora_inicio', '')
            ws['C7'].font = Font(name='Arial', size=10)
            ws['C7'].fill = info_fill
            ws['C7'].border = thin_border
            ws.merge_cells('C7')  

            # Fila 7: Hora de finalización
            ws['D7'].value = 'HORA FINALIZACIÓN:'
            ws['D7'].font = Font(bold=True, name='Arial', size=10)
            ws['D7'].fill = label_fill
            ws['D7'].border = thin_border
            ws.merge_cells('D7')  

            ws['E7'].value = actividad.get('hora_fin', '')
            ws['E7'].font = Font(name='Arial', size=10)
            ws['E7'].fill = info_fill
            ws['E7'].border = thin_border
            ws.merge_cells('E7:F7')  # Fusionar E7:F7

            # Ajustar inicio de la tabla de asistentes
            start_row = 10  # Dejar un espacio después de la información
            
            # Fila 5: Objetivo
            # Establecer etiqueta
            ws.merge_cells('G5:H7')
            ws['G5'].value = 'OBJETIVO:'
            ws['G5'].font = Font(bold=True, size=10, name='Arial')
            ws['G5'].fill = label_fill  # Color de fondo para la etiqueta
            # Aplicar borde completo a la celda de etiqueta
            for row in ws['G5:H7']:
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
            ws['G5'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Establecer valor del objetivo
            ws['I5'].value = actividad.get('objetivo', '').upper()
            ws['I5'].font = Font(name='Arial', size=10)
            ws['I5'].fill = info_fill
            # Aplicar borde completo a la celda de valor
            for row in ws['I5:K7']:
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
            ws['I5'].alignment = Alignment(wrap_text=True, vertical='top')
            
            # Fusionar celdas después de establecer valores
            ws.merge_cells('I5:K7')  # Objetivo ocupa 3 filas y 4 columnas (excluyendo la columna G con la etiqueta)
            
            # Ajustar altura de filas para el objetivo (usar auto_size para que se ajuste al contenido)
            ws.row_dimensions[5].auto_size = True
            ws.row_dimensions[6].auto_size = True
            ws.row_dimensions[7].auto_size = True
            
            # Fila inicial para la tabla de asistentes
            start_row = 8
            
        except Exception as e:
            logger.error(f"Error al generar el encabezado: {str(e)}")
            start_row = 1  # Si hay error, comenzar desde la primera fila
        
        # Definir todas las columnas posibles con sus etiquetas
        columnas_disponibles = [
            {'campo': 'numero', 'etiqueta': 'N°', 'visible_por_defecto': True},
            {'campo': 'fecha_registro', 'etiqueta': 'FECHA DE REGISTRO', 'visible_por_defecto': True},
            {'campo': 'nombre', 'etiqueta': 'NOMBRE COMPLETO', 'visible_por_defecto': True},
            {'campo': 'tipo_documento', 'etiqueta': 'TIPO DE DOCUMENTO', 'visible_por_defecto': True},
            {'campo': 'identificacion', 'etiqueta': 'NÚMERO DE DOCUMENTO', 'visible_por_defecto': True},
            {'campo': 'genero', 'etiqueta': 'GÉNERO', 'visible_por_defecto': True},
            {'campo': 'edad', 'etiqueta': 'EDAD', 'visible_por_defecto': True},
            {'campo': 'rango_edad', 'etiqueta': 'RANGO DE EDAD', 'visible_por_defecto': True},
            {'campo': 'comuna', 'etiqueta': 'COMUNA', 'visible_por_defecto': True},
            {'campo': 'barrio', 'etiqueta': 'BARRIO', 'visible_por_defecto': True},
            {'campo': 'telefono', 'etiqueta': 'TELÉFONO', 'visible_por_defecto': True},
            {'campo': 'correo', 'etiqueta': 'CORREO ELECTRÓNICO', 'visible_por_defecto': True},
            {'campo': 'estudia', 'etiqueta': '¿ESTUDIA?', 'visible_por_defecto': True},
            {'campo': 'nivel_educativo', 'etiqueta': 'NIVEL EDUCATIVO', 'visible_por_defecto': True},
            {'campo': 'sabe_leer', 'etiqueta': '¿LEE?', 'visible_por_defecto': True},
            {'campo': 'sabe_escribir', 'etiqueta': '¿ESCRIBE?', 'visible_por_defecto': True},
            {'campo': 'tipo_vivienda', 'etiqueta': 'TIPO DE VIVIENDA', 'visible_por_defecto': True},
            {'campo': 'situacion_laboral', 'etiqueta': '¿LABORA?', 'visible_por_defecto': True},
            {'campo': 'grupo_etnico', 'etiqueta': 'ÉTNIA', 'visible_por_defecto': True},
            {'campo': 'ayuda_humanitaria', 'etiqueta': '¿RECIBE AYUDA?', 'visible_por_defecto': True},
            {'campo': 'tipo_ayuda_humanitaria', 'etiqueta': 'TIPO DE AYUDA', 'visible_por_defecto': True},
            {'campo': 'discapacidad', 'etiqueta': '¿DISCAPACIDAD?', 'visible_por_defecto': True},
            {'campo': 'tipo_discapacidad', 'etiqueta': 'TIPO DE DISCAPACIDAD', 'visible_por_defecto': True},
            {'campo': 'nombre_cuidadora', 'etiqueta': 'NOMBRE DEL CUIDADOR/A', 'visible_por_defecto': True},
            {'campo': 'labora_actualmente', 'etiqueta': '¿TRABAJA?', 'visible_por_defecto': True},
            {'campo': 'victima_conflicto', 'etiqueta': '¿VÍCTIMA?', 'visible_por_defecto': True},
            {'campo': 'firma', 'etiqueta': 'FIRMA', 'visible_por_defecto': False}
        ]
        
        # Filtrar columnas según lo solicitado
        if columnas:
            # Filtrar solo las columnas que existen en columnas_disponibles
            columnas_seleccionadas = [col for col in columnas 
                                    if any(c['campo'] == col for c in columnas_disponibles)]
            
            # Si no hay columnas válidas, usar las visibles por defecto
            if not columnas_seleccionadas:
                columnas_seleccionadas = [c['campo'] for c in columnas_disponibles if c['visible_por_defecto']]
        else:
            # Si no se especifican columnas, usar las visibles por defecto
            columnas_seleccionadas = [c['campo'] for c in columnas_disponibles if c['visible_por_defecto']]
        
        # Obtener los encabezados para las columnas seleccionadas
        headers = []
        for col in columnas_disponibles:
            if col['campo'] in columnas_seleccionadas:
                headers.append(col['etiqueta'])
        
        # Agregar encabezados de la tabla
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[start_row].height = 60  # Ajusta este valor según necesites

        # Mapeo de campos a sus valores correspondientes
        def obtener_valor_campo(asistente, campo, indice, row_num=None):
            import time
            beneficiario = asistente.get('beneficiario', {})
            
            # Obtener fecha de registro del asistente o del beneficiario
            fecha_registro = asistente.get('fecha_asistencia') or asistente.get('fecha_registro')
            if not fecha_registro:
                fecha_registro = beneficiario.get('fecha_registro', '')
            
            # Formatear la fecha si existe
            if fecha_registro:
                try:
                    # Si es un string, intentar parsearlo
                    if isinstance(fecha_registro, str):
                        # Intentar diferentes formatos de fecha
                        formatos_fecha = [
                            '%Y-%m-%dT%H:%M:%S.%fZ',  # Formato con milisegundos y Z
                            '%Y-%m-%dT%H:%M:%S',       # Formato sin milisegundos
                            '%Y-%m-%d %H:%M:%S',        # Formato con espacio
                            '%Y-%m-%d',                 # Solo fecha
                            '%d/%m/%Y %H:%M:%S',        # Formato día/mes/año con hora
                            '%d/%m/%Y'                  # Solo fecha día/mes/año
                        ]
                        
                        fecha_obj = None
                        for formato in formatos_fecha:
                            try:
                                # Limpiar la cadena de fecha eliminando la parte de la zona horaria si existe
                                fecha_limpia = fecha_registro.split('+')[0].split('.')[0].strip()
                                fecha_obj = datetime.strptime(fecha_limpia, formato)
                                break
                            except (ValueError, AttributeError) as ve:
                                logger.debug(f"No se pudo parsear con formato {formato}: {str(ve)}")
                                continue
                        
                        if fecha_obj:
                            # Formatear como día/mes/año hora:minutos
                            fecha_registro = fecha_obj.strftime('%d/%m/%Y %H:%M')
                        else:
                            logger.warning(f"No se pudo formatear la fecha: {fecha_registro}")
                            # Intentar eliminar solo la parte de la zona horaria si existe
                            fecha_registro = str(fecha_registro).split('+')[0].split('.')[0].strip()
                    # Si ya es un objeto datetime
                    elif hasattr(fecha_registro, 'strftime'):
                        # Formatear como día/mes/año hora:minutos
                        fecha_registro = fecha_registro.strftime('%d/%m/%Y %H:%M')
                except Exception as e:
                    logger.warning(f"Error al formatear fecha {fecha_registro}: {str(e)}")
                    # Intentar eliminar solo la parte de la zona horaria si existe
                    fecha_registro = str(fecha_registro).split('+')[0].split('.')[0].strip()
            
            # Inicializar variable de firma
            firma_data = ''
            logger.info(f"Buscando firma para beneficiario: {beneficiario.get('nombre_completo', 'Sin nombre')}")
            logger.info(f"Campos del beneficiario: {list(beneficiario.keys())}")
            
            try:
                # Buscar firma digital
                if 'firma' in beneficiario and beneficiario['firma']:
                    logger.info("Firma digital encontrada en beneficiario")
                    firma_data = beneficiario['firma']
                    logger.info("Firma digital obtenida del beneficiario")
                
                # Buscar en el nivel superior del beneficiario
                if 'registro' in beneficiario and isinstance(beneficiario['registro'], dict):
                    if 'firma' in beneficiario['registro'] and beneficiario['registro']['firma']:
                        firma_data = beneficiario['registro']['firma']
                        logger.info("Firma digital encontrada en registro del beneficiario")
                
                logger.info(f"Firma digital obtenida para {beneficiario.get('nombre_completo', 'Sin nombre')}")
                
                # Procesar la firma digital si existe
                if firma_data and campo == 'firma' and row_num is not None:
                    try:
                        # Procesar la firma en formato base64
                        if 'base64,' in firma_data:
                            firma_data = firma_data.split('base64,')[1]
                        
                        # Decodificar la imagen
                        image_data = base64.b64decode(firma_data)
                        img = Image.open(BytesIO(image_data))
                        
                        # Redimensionar la imagen
                        max_size = (150, 50)
                        img.thumbnail(max_size, Image.Resampling.LANCZOS)
                        
                        # Guardar en buffer
                        img_byte_arr = BytesIO()
                        img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        
                        # Crear objeto de imagen para Excel
                        img_obj = OpenpyxlImage(img_byte_arr)
                        
                        # Calcular posición de la celda
                        col_idx = columnas_seleccionadas.index('firma') + 1
                        col_letter = get_column_letter(col_idx)
                        
                        # Insertar imagen en la celda
                        cell_reference = f"{col_letter}{row_num}"
                        ws.add_image(img_obj, cell_reference)
                        
                        # Ajustar altura de la fila para la imagen
                        ws.row_dimensions[row_num].height = 32
                        
                        # Ajustar ancho de la columna
                        ws.column_dimensions[col_letter].width = 20
                        
                        logger.info(f"Firma insertada en celda {cell_reference}")
                        
                    except Exception as e:
                        logger.error(f"Error al procesar la firma: {str(e)}", exc_info=True)
                        return "Error en firma"
                    
                    # Devolver cadena vacía ya que la imagen se insertó directamente
                    return ""
                
                # Si es el campo firma pero no hay datos
                if campo == 'firma':
                    return "Sin firma"

            except Exception as e:
                logger.warning(f"Error al obtener/enviar imagen QR: {str(e)}")
            
            # Obtener rango de edad
            rango_edad = beneficiario.get('rango_edad', '')
            if not rango_edad and 'edad' in beneficiario:
                try:
                    edad = int(float(beneficiario['edad']))  # Manejar tanto strings como números
                    if edad < 6:
                        rango_edad = "0-5"
                    elif edad < 12:
                        rango_edad = "6-11"
                    elif edad < 18:
                        rango_edad = "12-17"
                    elif edad < 26:
                        rango_edad = "18-25"
                    elif edad < 36:
                        rango_edad = "26-35"
                    elif edad < 46:
                        rango_edad = "36-45"
                    elif edad < 56:
                        rango_edad = "46-55"
                    elif edad < 66:
                        rango_edad = "56-65"
                    else:
                        rango_edad = "66+"
                except (ValueError, TypeError):
                    pass
            
            # Mapeo de campos a valores
            field_mapping = {
                'numero': str(indice + 1),
                'fecha_registro': fecha_registro,
                'nombre': beneficiario.get('nombre_completo', ''),
                'tipo_documento': beneficiario.get('tipo_documento', ''),
                'identificacion': beneficiario.get('numero_documento', ''),
                'genero': beneficiario.get('genero', ''),
                'edad': str(beneficiario.get('edad', '')),
                'rango_edad': rango_edad,
                'comuna': beneficiario.get('comuna', ''),
                'barrio': beneficiario.get('barrio', ''),
                 'telefono': beneficiario.get('numero_celular', ''),
                'correo': beneficiario.get('correo_electronico', ''),
                'estudia': 'Sí' if beneficiario.get('estudia_actualmente', False) else 'No',
                'nivel_educativo': beneficiario.get('nivel_educativo', ''),
                'sabe_leer': 'Sí' if beneficiario.get('sabe_leer', False) else 'No',
                'sabe_escribir': 'Sí' if beneficiario.get('sabe_escribir', False) else 'No',
                'tipo_vivienda': beneficiario.get('tipo_vivienda', ''),
                'situacion_laboral': beneficiario.get('situacion_laboral', ''),
                  'grupo_etnico': beneficiario.get('etnia', ''),
                'ayuda_humanitaria': 'Sí' if beneficiario.get('ayuda_humanitaria', False) else 'No',
                'tipo_ayuda_humanitaria': beneficiario.get('descripcion_ayuda_humanitaria', ''),
                'discapacidad': 'Sí' if beneficiario.get('tiene_discapacidad', False) else 'No',
                'tipo_discapacidad': beneficiario.get('tipo_discapacidad', ''),
                'nombre_cuidadora': beneficiario.get('nombre_cuidadora', ''),
                'labora_actualmente': 'Sí' if beneficiario.get('labora_actualmente') else 'No',
                'victima_conflicto': 'Sí' if beneficiario.get('victima_conflicto') else 'No',
                'firma': beneficiario.get('firma', '')  # Agregar la firma al mapeo de campos
            }
            
            return field_mapping.get(campo, '')
        
        # Llenar datos - Comenzar después del encabezado
        start_row_table = start_row + 1  # Fila después del encabezado de la tabla
        
        for i, asistente in enumerate(actividad.get('asistentes', [])):
            try:
                # Obtener valores para cada columna seleccionada
                row_data = []
                # El número de fila es start_row_table (fila inicial de datos) + i (índice del asistente)
                # start_row ya incluye el encabezado, por eso no sumamos 1 adicional
                row_num = start_row_table + i
                logger.info(f"Procesando fila {row_num} para asistente {i}")
                
                for col in columnas_disponibles:
                    if col['campo'] in columnas_seleccionadas:
                        valor = obtener_valor_campo(asistente, col['campo'], i, row_num)
                        row_data.append(valor)
                
                # Agregar fila a la hoja
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=start_row_table + i, column=col_num, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
            except Exception as e:
                logger.error(f"Error al procesar asistente {i}: {str(e)}", exc_info=True)
                continue
        
        # Ajustar ancho de columnas (solo para las columnas de datos, no las del encabezado)
        for col_num in range(1, len(columnas_seleccionadas) + 1):
            column_letter = get_column_letter(col_num)
            
            # Establecer ancho fijo para la columna del nombre completo
            if columnas_seleccionadas[col_num-1] == 'nombre':
                ws.column_dimensions[column_letter].width = 30  # Ancho fijo para la columna de nombre
                continue
                
            # Solo ajustar columnas que no están en el encabezado combinado
            if column_letter in ['C', 'D', 'E']:  # Columnas del título combinado
                ws.column_dimensions[column_letter].width = 20  # Ancho fijo
                continue
                
            # Para otras columnas, ajustar automáticamente el ancho
            max_length = 0
            for row in range(start_row_table, start_row_table + len(actividad.get('asistentes', []))):
                try:
                    cell_value = ws.cell(row=row, column=col_num).value
                    if cell_value and len(str(cell_value)) > max_length:
                        max_length = len(str(cell_value))
                except:
                    pass
            
            if max_length > 0:
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Máximo 50 caracteres
        
        # Crear un archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            excel_path = tmp_file.name
        
        # Guardar el libro de Excel
        wb.save(excel_path)
        logger.info(f"Archivo Excel generado: {excel_path}")
        
        # Enviar el archivo como respuesta
        return send_file(
            excel_path,  # Usar excel_path en lugar de temp_filename
            as_attachment=True,
            download_name=f"asistencia_{actividad.get('tema', 'actividad')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error en exportar_actividad: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': 'Error al exportar la actividad',
            'error': str(e)
        }), 500

def subir_logo():
    try:
        from app.routes.auth import token_required
        
        if 'logo' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No se ha subido ningún archivo'
            }), 400
            
        file = request.files['logo']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': 'No se ha seleccionado ningún archivo'
            }), 400
            
        # Validar extensión
        allowed_extensions = {'png', 'jpg', 'jpeg', 'svg'}
        filename = secure_filename(file.filename)
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return jsonify({
                'success': False,
                'message': 'Extensión de archivo no permitida. Use PNG, JPG, JPEG o SVG.'
            }), 400
            
        # Validar tamaño (máx 2MB)
        if file.content_length > 2 * 1024 * 1024:
            return jsonify({
                'success': False,
                'message': 'El archivo es demasiado grande (máx 2MB)'
            }), 400
            
        # Crear directorio de uploads si no existe
        upload_folder = os.path.join(current_app.root_path, '..', '..', 'uploads')
        os.makedirs(upload_folder, exist_ok=True)
        
        # Generar nombre único para el archivo
        unique_filename = f"logo_{int(time.time())}_{filename}"
        filepath = os.path.join(upload_folder, unique_filename)
        
        # Guardar el archivo
        file.save(filepath)
        
        # Devolver la ruta relativa con formato URL
        rel_path = os.path.join('uploads', unique_filename).replace('\\', '/')
        
        return jsonify({
            'success': True,
            'url': f'/{rel_path}',
            'message': 'Logo subido correctamente'
        })
        
    except Exception as e:
        logger.error(f"Error al subir logo: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Error al exportar la reunión',
            'error': str(e)
        }), 500

def exportar_reunion(reunion_id, columnas=None):
    """
    Exporta la lista de asistentes de una reunión a un archivo Excel.
    
    Args:
        reunion_id (str): ID de la reunión a exportar
        columnas (list, optional): Lista de columnas a incluir en el reporte. 
                                 Si es None, se incluirán las columnas por defecto.
    
    Returns:
        Response: Archivo Excel para descargar o mensaje de error
    """
    try:
        from app.models.actividad import ActividadModel
        from app.models.asistente import AsistenteModel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from flask import jsonify, send_file, current_app
        import tempfile
        import os
        import logging
        from bson import ObjectId
        from datetime import datetime
        import traceback
        import json
        import base64
        from io import BytesIO
        from PIL import Image
        from openpyxl.drawing.image import Image as OpenpyxlImage
                # Configurar logger
        logger = logging.getLogger(__name__)
        logger.info(f"Iniciando exportación de reunión a Excel: {reunion_id}")
        
        # Obtener la reunión de la base de datos
        actividad_model = ActividadModel()
        reunion = actividad_model.obtener_actividad_por_id(reunion_id)
        
        if not reunion or reunion.get('tipo') != 'reunion':  # Cambiado 'reunión' a 'reunion' (sin tilde)
            logger.error(f"Reunión no encontrada o no es una reunión: {reunion_id}")
            return jsonify({
                'success': False,
                'message': 'Reunión no encontrada o no es una reunión',
                'error': f'No se encontró la reunión con ID {reunion_id} o no es una reunión válida. Tipo encontrado: {reunion.get("tipo") if reunion else "Ninguno"}'
            }), 404
            
        # Obtener los asistentes de la reunión
        asistentes_reunion = reunion.get('asistentes', [])
        
        if not asistentes_reunion:
            logger.error("No hay asistentes en la reunión")
            return jsonify({
                'success': False,
                'message': 'No hay asistentes para exportar',
                'error': 'No hay datos para exportar'
            }), 404
            
        # Obtener datos completos de los asistentes
        asistente_model = AsistenteModel(current_app.config['db'])
        asistentes_completos = []
        
        for i, asistente_ref in enumerate(asistentes_reunion, 1):
            try:
                logger.info(f"Procesando asistente_ref: {json.dumps(asistente_ref, default=str)}")
                # Obtener los datos del asistente
                asistente = {
                    'numero': i,
                    'nombre': asistente_ref.get('nombre', ''),
                    'cedula': asistente_ref.get('cedula', ''),
                    'dependencia': asistente_ref.get('dependencia', ''),
                    'cargo': asistente_ref.get('cargo', ''),
                    'tipo_participacion': asistente_ref.get('tipo_participacion', ''),
                    'telefono': asistente_ref.get('telefono', ''),
                    'email': asistente_ref.get('email', ''),
                    'firma': asistente_ref.get('firma', '')  # Guardamos la firma real
                }
                logger.info(f"Antes de beneficiario - asistente: {json.dumps(asistente, default=str)}")

                # Si hay un beneficiario_id, obtener sus datos adicionales
                
                # Busca esta sección en tu código (alrededor de la línea 1400)
                if 'beneficiario_id' in asistente_ref and asistente_ref['beneficiario_id']:
                    try:
                        beneficiario = asistente_model.obtener_por_id(asistente_ref['beneficiario_id'])
                        logger.info(f"Beneficiario encontrado: {json.dumps(beneficiario, default=str) if beneficiario else 'No encontrado'}")
                        
                        if beneficiario:
                            # Actualizar con datos del beneficiario
                            asistente.update({
                                'nombre': beneficiario.get('nombre', beneficiario.get('nombre_completo', asistente['nombre'])),
                                'cedula': beneficiario.get('cedula', asistente['cedula']),
                                'dependencia': beneficiario.get('dependencia', asistente['dependencia']),
                                'cargo': beneficiario.get('cargo', asistente['cargo']),
                                'tipo_participacion': beneficiario.get('tipo_participacion', asistente['tipo_participacion']),
                                'telefono': beneficiario.get('telefono', asistente['telefono']),
                                'email': beneficiario.get('email', asistente['email']),
                                'firma': beneficiario.get('firma', asistente['firma'])
                            })
                            logger.info(f"Después de actualizar con beneficiario: {json.dumps(asistente, default=str)}")
                    except Exception as e:
                        logger.error(f"Error al obtener datos del beneficiario {asistente_ref['beneficiario_id']}: {str(e)}", exc_info=True)
                logger.info(f"Asistente final a agregar: {json.dumps(asistente, default=str)}")
                asistentes_completos.append(asistente)
            except Exception as e:
                logger.error(f"Error al procesar asistente: {str(e)}")
                continue
        
        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia Reunión"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Color de fondo para celdas de información (blanco)
        info_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Color de fondo para celdas de etiquetas (gris claro)
        label_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Agregar encabezado con logo e información de la actividad
        try:
            # Ruta al logo (usando ruta absoluta)
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            logo_path = os.path.join(base_dir, 'frontend', 'src', 'fondo', 'logo.png')
                       
            # Configuración de estilos
            font_size = 10
            font_name = 'Arial'
            
            # Configurar dimensiones de celdas
            ws.column_dimensions['A'].width = 15  # Columna A para el logo
            ws.column_dimensions['B'].width = 15  # Columna B para etiquetas
            ws.column_dimensions['C'].width = 20  # Columna C para valores
            ws.column_dimensions['D'].width = 15  # Columna D
            ws.column_dimensions['E'].width = 15  # Columna E
            ws.column_dimensions['F'].width = 20  # Columna F
            ws.column_dimensions['G'].width = 15  # Columna G
            ws.column_dimensions['H'].width = 15  # Columna H
            ws.column_dimensions['I'].width = 15  # Columna I
            ws.column_dimensions['J'].width = 15  # Columna J
            ws.column_dimensions['K'].width = 20  # Columna K
            
            # Altura de filas (se deja la altura por defecto)
            # No establecemos altura fija para que use la altura por defecto
            
            # Agregar logo (ocupa A1:C4)
            logo_size = 300  # Tamaño del logo más grande
            
            if os.path.exists(logo_path):
                try:
                    img = XLImage(logo_path)
                    # Ajustar tamaño manteniendo la relación de aspecto
                    img_ratio = img.height / img.width
                    # Redimensionar manteniendo la relación de aspecto
                    img.width = logo_size
                    img.height = int(logo_size * img_ratio)
                    
                    # No ajustamos la altura de las filas, se mantiene la altura por defecto
                    
                    # Agregar el logo
                    ws.add_image(img, 'A1')
                    logger.info("Logo agregado exitosamente")
                    
                    # Fusionar celdas para el logo (4 filas de alto x 3 columnas de ancho)
                    ws.merge_cells('A1:C4')
                    
                except Exception as e:
                    logger.error(f"Error al cargar la imagen: {str(e)}")
            else:
                logger.warning(f"No se encontró el archivo de logo en: {logo_path}")
            
            # Título principal (D1:K1)
            ws.merge_cells('D1:I1')
            title_cell = ws['D1']
            title_cell.value = 'FORMATO REGISTRO DE ASISTENCIA'
            title_cell.font = Font(bold=True, size=14, name='Arial')
            title_cell.fill = info_fill
            # title_cell.border = thin_border
            for row in ws['D1:I1']:
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Subtítulo (D2:K2)
            ws.merge_cells('D2:I2')
            subtitle_cell = ws['D2']
            subtitle_cell.value = 'ALCALDÍA MUNICIPAL DE QUIBDÓ'
            subtitle_cell.font = Font(bold=True, size=12, name='Arial')
            subtitle_cell.fill = info_fill
            for row in ws['D2:I2']:
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
            subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
                        

            # Fila 4: Dependencia
            ws['D3'].value = 'DEPENDENCIA:'
            ws['D3'].font = Font(bold=True, name='Arial', size=10)
            ws['D3'].fill = label_fill
            ws['D3'].border = thin_border
            ws.merge_cells('D3:E3')  # Fusionar D4:E4

            ws['F3'].value = reunion.get('dependencia', '').upper()
            ws['F3'].font = Font(name='Arial', size=10)
            ws['F3'].fill = info_fill
            ws['F3'].border = thin_border
            ws.merge_cells('F3:I3')  # Fusionar F4:I4

               # Fila 4: Dependencia
      
            # Combinar celdas sin texto
            ws.merge_cells('D4:I4')  # Fusionar D4:K4
            ws['D4'].fill = label_fill
            ws['D4'].border = thin_border

            # Fila 5: Tema
            ws['A5'].value = 'TEMA:'
            ws['A5'].font = Font(bold=True, name='Arial', size=10)
            ws['A5'].fill = label_fill
            ws['A5'].border = thin_border
            ws.merge_cells('A5:B5') 

            ws['C5'].value = reunion.get('tema', '').upper()
            ws['C5'].font = Font(name='Arial', size=10)
            ws['C5'].fill = info_fill
            ws['C5'].border = thin_border
            ws.merge_cells('C5:f5')  # Fusionar C5:F5

            # Fila 6: Fecha
            ws['A6'].value = 'FECHA:'
            ws['A6'].font = Font(bold=True, name='Arial', size=10)
            ws['A6'].fill = label_fill
            ws['A6'].border = thin_border
            ws.merge_cells('A6:B6')

            if 'fecha' in reunion:
                fecha_obj = datetime.strptime(reunion['fecha'], '%Y-%m-%dT%H:%M:%S')  # Si viene como string
                # O si ya es un objeto datetime: fecha_obj = reunion['fecha']
                fecha_formateada = fecha_obj.strftime('%d/%m/%Y')  # Formato DD/MM/YYYY
                ws['C6'].value = f"FECHA: {fecha_formateada}"
                ws['C6'].font = Font(name='Arial', size=10)
                ws['C6'].fill = info_fill
                ws['C6'].border = thin_border
            ws.merge_cells('C6') 

            # Fila 6: Lugar
            ws['D6'].value = 'LUGAR:'
            ws['D6'].font = Font(bold=True, name='Arial', size=10)
            ws['D6'].fill = label_fill
            ws['D6'].border = thin_border
            ws.merge_cells('D6') 

            ws['E6'].value = reunion.get('lugar', '').upper()
            ws['E6'].font = Font(name='Arial', size=10)
            ws['E6'].fill = info_fill
            ws['E6'].border = thin_border
            ws.merge_cells('E6:F6')  # Fusionar E6:F6

            # Fila 7: Hora de inicio
            ws['A7'].value = 'HORA INICIO:'
            ws['A7'].font = Font(bold=True, name='Arial', size=10)
            ws['A7'].fill = label_fill
            ws['A7'].border = thin_border
            ws.merge_cells('A7:B7') 

            ws['C7'].value = reunion.get('hora_inicio', '')
            ws['C7'].font = Font(name='Arial', size=10)
            ws['C7'].fill = info_fill
            ws['C7'].border = thin_border
            ws.merge_cells('C7')

            # Fila 7: Hora de finalización
            ws['D7'].value = 'HORA FINALIZACIÓN:'
            ws['D7'].font = Font(bold=True, name='Arial', size=10)
            ws['D7'].fill = label_fill
            ws['D7'].border = thin_border
            ws.merge_cells('D7')  


            
            ws['E7'].value = reunion.get('hora_fin', '')
            ws['E7'].font = Font(name='Arial', size=10)
            ws['E7'].fill = info_fill
            ws['E7'].border = thin_border
            ws.merge_cells('E7:F7')  # Fusionar E7:F7
            
             # Establecer etiqueta
            ws.merge_cells('G5:G7')
            ws['G5'].value = 'OBJETIVO:'
            ws['G5'].font = Font(bold=True, size=10, name='Arial')
            ws['G5'].fill = label_fill  # Color de fondo para la etiqueta
            ws['G5'].border = thin_border
          
            for row in ws['G5:G7']:
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
            ws['G5'].alignment = Alignment(horizontal='center', vertical='center')
                        
            # Establecer valor del objetivo
            ws['H5'].value = reunion.get('objetivo', '').upper()
            ws['H5'].font = Font(name='Arial', size=10)
            ws['H5'].fill = info_fill
            # Aplicar borde completo a la celda de valor
            for row in ws['H5:I7']:
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
            ws['H5'].alignment = Alignment(wrap_text=True, vertical='top')
            
            # Fusionar celdas después de establecer valores
            ws.merge_cells('H5:I7')  # Objetivo ocupa 3 filas y 4 columnas (excluyendo la columna G con la etiqueta)
            
            # Ajustar altura de filas para el objetivo (usar auto_size para que se ajuste al contenido)
            ws.row_dimensions[5].auto_size = True
            ws.row_dimensions[6].auto_size = True
            ws.row_dimensions[7].auto_size = True
            
            # Fila inicial para la tabla de asistentes
           
            
            # Agregar encabezados de la tabla de asistentes
            # En la función exportar_reunion, modificar la sección de encabezados:
           # En la sección de encabezados, actualizar a:
           # En la sección de encabezados
            headers = ['NOMBRE COMPLETO', '', 'CÉDULA', 'DEPENDENCIA', 'CARGO', 'TIPO PARTICIPACIÓN', 'TELÉFONO', 'EMAIL', 'FIRMA']
            start_row = 8

            # Escribir encabezados
            for col_num, header in enumerate(headers, 1):  # Comenzar desde columna 1
                cell = ws.cell(row=start_row, column=col_num, value=header if header else '')
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Combinar celdas del encabezado para NOMBRE COMPLETO (A y B)
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)

            # Ajustar el ancho de las columnas
            column_widths = {
                1: 14,  # Primera parte de NOMBRE (A)
                2: 14,  # Segunda parte de NOMBRE (B) - se combinará con A
                3: 18,  # CÉDULA (C)
                4: 25,  # DEPENDENCIA (D)
                5: 25,  # CARGO (E)
                6: 20,  # TIPO PARTICIPACIÓN (F)
                7: 13,  # TELÉFONO (G)
                8: 35,  # EMAIL (H)
                9: 20   # FIRMA (I)
            }

            for col_num, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col_num)].width = width

            # Escribir datos de asistentes
            for row_num, asistente in enumerate(asistentes_completos, start_row + 1):
                try:
                    # NOMBRE COMPLETO (columnas A y B combinadas)
                    nombre = asistente.get('nombre', '')
                    ws.cell(row=row_num, column=1, value=nombre).border = thin_border
                    ws.cell(row=row_num, column=2, value='').border = thin_border
                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
                    
                    # CÉDULA (columna C)
                    ws.cell(row=row_num, column=3, value=asistente.get('cedula', '')).border = thin_border
                    
                    # DEPENDENCIA (columna D)
                    dependencia = f"{asistente.get('dependencia', '')} {asistente.get('dependencia_adicional', '')}".strip()
                    ws.cell(row=row_num, column=4, value=dependencia).border = thin_border
                    
                    # CARGO (columna E)
                    ws.cell(row=row_num, column=5, value=asistente.get('cargo', '')).border = thin_border
                    
                    # TIPO PARTICIPACIÓN (columna F)
                    ws.cell(row=row_num, column=6, value=asistente.get('tipo_participacion', '')).border = thin_border
                    
                    # TELÉFONO (columna G)
                    ws.cell(row=row_num, column=7, value=asistente.get('telefono', '')).border = thin_border
                    
                    # EMAIL (columna H)
                    ws.cell(row=row_num, column=8, value=asistente.get('email', '')).border = thin_border
                    
                    # FIRMA (columna I)
                    if 'firma' in asistente and asistente['firma']:
                        try:
                            img_data = asistente['firma']
                            if 'base64,' in img_data:
                                img_data = img_data.split('base64,')[1]
                            
                            image_data = base64.b64decode(img_data)
                            img = Image.open(BytesIO(image_data))
                            
                            max_size = (150, 50)
                            img.thumbnail(max_size, Image.Resampling.LANCZOS)
                            
                            img_byte_arr = BytesIO()
                            img.save(img_byte_arr, format='PNG')
                            img_byte_arr.seek(0)
                            
                            img_obj = OpenpyxlImage(img_byte_arr)
                            ws.add_image(img_obj, f'I{row_num}')  # Columna I para la firma
                            ws.row_dimensions[row_num].height = 32
                            
                        except Exception as e:
                            logger.error(f"Error al procesar firma: {str(e)}")
                            ws.cell(row=row_num, column=9, value="Error en firma").border = thin_border
                    else:
                        ws.cell(row=row_num, column=9, value="Sin firma").border = thin_border
                        
                except Exception as e:
                    logger.error(f"Error al procesar asistente: {str(e)}")
                    continue
            # Ajustar el ancho de las columnas al contenido
            for col_num in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                # Saltar si la columna ya tiene un ancho definido
                if column_letter in ws.column_dimensions and ws.column_dimensions[column_letter].width is not None:
                    continue
                    
                for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                    cell = row[0]
                    # Verificar si la celda es parte de un rango combinado
                    is_merged = False
                    for merge_range in ws.merged_cells.ranges:
                        if cell.coordinate in merge_range:
                            is_merged = True
                            break
                    
                    if not is_merged and cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                
                if max_length > 0:
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = min(adjusted_width, 30)  # Máximo ancho de 30
            
            # Crear archivo temporal
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name
            
            # Enviar el archivo
            return send_file(
                tmp_path,
                as_attachment=True,
                download_name=f"asistencia_reunion_{reunion_id}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            logger.error(f"Error al generar el archivo Excel: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({
                'success': False,
                'message': 'Error al generar el archivo Excel',
                'error': str(e)
            }), 500
            
    except Exception as e:
        logger.error(f"Error en exportar_reunion: {str(e)}")
        # Definir todas las columnas posibles con sus etiquetas
        columnas_disponibles = [
    
            {'campo': 'nombre', 'etiqueta': 'NOMBRE COMPLETO', 'visible_por_defecto': True},
            {'campo': 'cedula', 'etiqueta': 'CÉDULA', 'visible_por_defecto': True},
            {'campo': 'dependencia', 'etiqueta': 'DEPENDENCIA', 'visible_por_defecto': True},
            {'campo': 'cargo', 'etiqueta': 'CARGO', 'visible_por_defecto': True},
            {'campo': 'tipo_participacion', 'etiqueta': 'TIPO PARTICIPACIÓN', 'visible_por_defecto': True},
            {'campo': 'telefono', 'etiqueta': 'TELÉFONO', 'visible_por_defecto': True},
            {'campo': 'email', 'etiqueta': 'CORREO ELECTRÓNICO', 'visible_por_defecto': True},
            {'campo': 'firma', 'etiqueta': 'FIRMA', 'visible_por_defecto': True}
        ]

        # Filtrar columnas según lo solicitado
        if columnas:
            columnas_seleccionadas = [col for col in columnas if any(c['campo'] == col for c in columnas_disponibles)]
            if not columnas_seleccionadas:
                columnas_seleccionadas = [c['campo'] for c in columnas_disponibles if c['visible_por_defecto']]
        else:
            columnas_seleccionadas = [c['campo'] for c in columnas_disponibles if c['visible_por_defecto']]

        # Obtener los encabezados para las columnas seleccionadas
        headers = [col['etiqueta'] for col in columnas_disponibles if col['campo'] in columnas_seleccionadas]

        # Agregar encabezados de la tabla
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Función auxiliar para obtener el valor de un campo de un asistente
        def obtener_valor_campo(asistente, campo, indice, row_num=None):
            try:
                # Mapeo de campos a valores directamente del objeto asistente
                field_mapping = {
                  
                    'nombre': asistente.get('nombre', '').strip(),
                    'cedula': asistente.get('cedula', '').strip(),
                    'dependencia': asistente.get('dependencia', '').strip(),
                    'cargo': asistente.get('cargo', '').strip(),
                    'tipo_participacion': asistente.get('tipo_participacion', '').strip(),
                    'telefono': asistente.get('telefono', '').strip(),
                    'email': asistente.get('email', '').strip(),
                    'firma': asistente.get('firma', '')  # No aplicamos strip() a la firma
                }
                
                valor = field_mapping.get(campo, '')
               
 
                # Manejo especial para el campo firma
                if campo == 'firma':
                    if valor and isinstance(valor, str) and valor.startswith('data:image/png;base64,'):
                        return 'Firmado'
                    else:
                        return 'Pendiente'
                        
                return valor
                
            except Exception as e:
                logger.error(f"Error al obtener valor del campo {campo}: {str(e)}")
                return ''

        # Llenar datos - Comenzar después del encabezado
        start_row_table = start_row + 1  # Fila después del encabezado de la tabla
        
        # Log de depuración
        logger.info(f"Asistentes completos a exportar: {json.dumps(asistentes_completos, default=str, indent=2)}")
        logger.info(f"Columnas seleccionadas: {columnas_seleccionadas}")
        
        for i, asistente in enumerate(asistentes_completos):
            try:
                # Obtener valores para cada columna seleccionada
                row_data = []
                row_num = start_row_table + i
                
                logger.info(f"Procesando asistente {i+1}: {asistente.get('nombre', 'Sin nombre')}")
                logger.info(f"Datos del asistente: {json.dumps(asistente, default=str)}")
                
                for col in columnas_disponibles:
                    if col['campo'] in columnas_seleccionadas:
                        valor = obtener_valor_campo(asistente, col['campo'], i, row_num)
                        logger.info(f"  - Columna '{col['campo']}': {valor}")
                        row_data.append(valor)
                
                # Agregar fila a la hoja
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                logger.info(f"Fila {row_num} procesada correctamente")
                
                
            except Exception as e:
                logger.error(f"Error al procesar asistente {i}: {str(e)}", exc_info=True)
                continue
        
        # Ajustar ancho de columnas (solo para las columnas de datos, no las del encabezado)
        for col_num in range(1, len(columnas_seleccionadas) + 1):
            column_letter = get_column_letter(col_num)
            
            # Establecer ancho fijo para la columna del nombre completo
            if columnas_seleccionadas[col_num-1] == 'nombre':
                ws.column_dimensions[column_letter].width = 30  # Ancho fijo para la columna de nombre
                continue
                
            # Solo ajustar columnas que no están en el encabezado combinado
            if column_letter in ['C', 'D', 'E']:  # Columnas del título combinado
                ws.column_dimensions[column_letter].width = 20  # Ancho fijo
                continue
                
            # Para otras columnas, ajustar automáticamente el ancho
            max_length = 0
            for row in range(start_row_table, start_row_table + len(actividad.get('asistentes', []))):
                try:
                    cell_value = ws.cell(row=row, column=col_num).value
                    if cell_value and len(str(cell_value)) > max_length:
                        max_length = len(str(cell_value))
                except:
                    pass
            
            if max_length > 0:
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Máximo 50 caracteres
        
        # Crear un archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            excel_path = tmp_file.name
        
        # Guardar el libro de Excel
        wb.save(excel_path)
        logger.info(f"Archivo Excel generado: {excel_path}")
        
        # Enviar el archivo como respuesta
        return send_file(
            excel_path,  # Usar excel_path en lugar de temp_filename
            as_attachment=True,
            download_name=f"asistencia_{actividad.get('tema', 'actividad')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error en exportar_actividad: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': 'Error al exportar la actividad',
            'error': str(e)
        }), 500