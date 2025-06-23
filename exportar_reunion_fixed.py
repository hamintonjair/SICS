def exportar_reunion(reunion_id, columnas=None):
    try:
        from app.models.actividad import ActividadModel
        from app.models.beneficiario import BeneficiarioModel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        import tempfile
        import os
        import requests
        import qrcode
        from io import BytesIO
        from bson import ObjectId
        from datetime import datetime
        from PIL import Image as PILImage
        from PIL import ImageDraw, ImageFont
        import logging
        
        logger = logging.getLogger(__name__)
        
        logger.info(f"Iniciando exportación de reunión a Excel: {reunion_id}")
        logger.info(f"Columnas seleccionadas: {columnas}")
        
        # Crear un libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"
        
        # Obtener la reunión de la base de datos
        actividad_model = ActividadModel()
        reunion = actividad_model.obtener_actividad_por_id(reunion_id)
        
        if not reunion:
            logger.error(f"Reunión no encontrada: {reunion_id}")
            return {
                'success': False,
                'message': 'Reunión no encontrada',
                'error': f'No se encontró la reunión con ID {reunion_id}'
            }, 404
        
        logger.info(f"Reunión encontrada: {reunion.get('tema', 'Sin nombre')}")
        
        # Obtener los datos completos de los asistentes
        if 'asistentes' in reunion and reunion['asistentes']:
            beneficiario_model = BeneficiarioModel()
            for asistente in reunion['asistentes']:
                if 'beneficiario_id' in asistente and asistente['beneficiario_id']:
                    try:
                        logger.info(f"Buscando asistente con ID: {asistente['beneficiario_id']}")
                        beneficiario = beneficiario_model.obtener_beneficiario_por_id(asistente['beneficiario_id'])
                        if beneficiario:
                            # Actualizar los datos del asistente con la información del beneficiario
                            asistente.update({
                                'nombre': beneficiario.get('nombre_completo', ''),
                                'cedula': beneficiario.get('numero_documento', ''),
                                'dependencia': reunion.get('dependencia', ''),
                                'telefono': beneficiario.get('numero_celular', ''),
                                'email': beneficiario.get('correo_electronico', ''),
                                'genero': beneficiario.get('genero', ''),
                                'barrio': beneficiario.get('barrio', ''),
                                'comuna': beneficiario.get('comuna', '')
                            })
                        else:
                            logger.warning(f"No se encontró el beneficiario con ID: {asistente['beneficiario_id']}")
                    except Exception as e:
                        logger.error(f"Error al obtener beneficiario {asistente['beneficiario_id']}: {str(e)}", exc_info=True)
                        continue
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Color de fondo para celdas de información (blanco)
        info_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Color de fondo para celdas de etiquetas (gris claro)
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Agregar encabezado con logo e información de la reunión
        try:
            # Ruta al logo (usando ruta absoluta)
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            logo_path = os.path.join(base_dir, 'frontend', 'src', 'fondo', 'logo.png')
            logger.info(f"Ruta base del proyecto: {base_dir}")
            logger.info(f"Ruta completa del logo: {logo_path}")
            logger.info(f"El archivo existe: {os.path.exists(logo_path)}")
            
            # Configuración de estilos
            font_size = 10
            font_name = 'Arial'
            
            # Configurar dimensiones de celdas
            ws.column_dimensions['A'].width = 15  # Columna A para el logo
            ws.column_dimensions['B'].width = 15  # Columna B para etiquetas
            ws.column_dimensions['C'].width = 20  # Columna C para valores
            ws.column_dimensions['D'].width = 15  # Columna D
            ws.column_dimensions['E'].width = 15  # Columna E
            ws.column_dimensions['F'].width = 20  # Columna F
            ws.column_dimensions['G'].width = 15  # Columna G
            ws.column_dimensions['H'].width = 15  # Columna H
            ws.column_dimensions['I'].width = 15  # Columna I
            ws.column_dimensions['J'].width = 15  # Columna J
            ws.column_dimensions['K'].width = 20  # Columna K
            
            # Agregar logo (ocupa A1:C4)
            if os.path.exists(logo_path):
                try:
                    img = XLImage(logo_path)
                    # Ajustar tamaño manteniendo la relación de aspecto
                    img_ratio = img.height / img.width
                    img.width = 100  # Tamaño más pequeño para el logo
                    img.height = int(100 * img_ratio)
                    
                    # Agregar el logo
                    ws.add_image(img, 'A1')
                    logger.info("Logo agregado exitosamente")
                    
                    # Fusionar celdas para el logo
                    ws.merge_cells('A1:C4')
                    
                except Exception as e:
                    logger.error(f"Error al cargar la imagen: {str(e)}")
            else:
                logger.warning(f"No se encontró el archivo de logo en: {logo_path}")
            
            # Título principal (D1:K1)
            ws.merge_cells('D1:K1')
            title_cell = ws['D1']
            title_cell.value = 'FORMATO REGISTRO DE ASISTENCIA A REUNIÓN'
            title_cell.font = Font(bold=True, size=14, name='Arial')
            title_cell.fill = info_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Subtítulo (D2:K2)
            ws.merge_cells('D2:K2')
            subtitle_cell = ws['D2']
            subtitle_cell.value = 'ALCALDÍA MUNICIPAL DE QUIBDÓ'
            subtitle_cell.font = Font(bold=True, size=12, name='Arial')
            subtitle_cell.fill = info_fill
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Información de la reunión
            # Fila 3: Dependencia
            ws['D3'].value = 'DEPENDENCIA:'
            ws['D3'].font = Font(bold=True, name='Arial', size=10)
            ws['D3'].fill = label_fill
            ws.merge_cells('D3:E3')
            
            ws['F3'].value = reunion.get('dependencia', '').upper()
            ws['F3'].font = Font(name='Arial', size=10)
            ws['F3'].fill = info_fill
            ws.merge_cells('F3:K3')
            
            # Fila 4: Tema
            ws['D4'].value = 'TEMA:'
            ws['D4'].font = Font(bold=True, name='Arial', size=10)
            ws['D4'].fill = label_fill
            ws.merge_cells('D4:E4')
            
            ws['F4'].value = reunion.get('tema', '').upper()
            ws['F4'].font = Font(name='Arial', size=10)
            ws['F4'].fill = info_fill
            ws.merge_cells('F4:K4')
            
            # Fila 5: Fecha
            ws['D5'].value = 'FECHA:'
            ws['D5'].font = Font(bold=True, name='Arial', size=10)
            ws['D5'].fill = label_fill
            ws.merge_cells('D5:E5')
            
            fecha = reunion.get('fecha', '')
            if isinstance(fecha, datetime):
                fecha = fecha.strftime('%Y-%m-%d')
            ws['F5'].value = fecha
            ws['F5'].font = Font(name='Arial', size=10)
            ws['F5'].fill = info_fill
            ws.merge_cells('F5:G5')
            
            # Fila 5: Hora
            ws['H5'].value = 'HORA:'
            ws['H5'].font = Font(bold=True, name='Arial', size=10)
            ws['H5'].fill = label_fill
            ws.merge_cells('H5:I5')
            
            hora_inicio = reunion.get('hora_inicio', '')
            if isinstance(hora_inicio, str) and 'T' in hora_inicio:
                hora_inicio = hora_inicio.split('T')[1][:5]  # Extraer solo la hora
            ws['J5'].value = hora_inicio
            ws['J5'].font = Font(name='Arial', size=10)
            ws['J5'].fill = info_fill
            ws.merge_cells('J5:K5')
            
            # Fila 6: Lugar
            ws['D6'].value = 'LUGAR:'
            ws['D6'].font = Font(bold=True, name='Arial', size=10)
            ws['D6'].fill = label_fill
            ws.merge_cells('D6:E6')
            
            ws['F6'].value = reunion.get('lugar', '').upper()
            ws['F6'].font = Font(name='Arial', size=10)
            ws['F6'].fill = info_fill
            ws.merge_cells('F6:K6')
            
            # Ajustar inicio de la tabla de asistentes
            start_row = 8
            
        except Exception as e:
            logger.error(f"Error al generar el encabezado: {str(e)}")
            start_row = 1  # Si hay error, comenzar desde la primera fila
        
        # Definir columnas para la reunión
        columnas_disponibles = [
            {'campo': 'numero', 'etiqueta': 'N°', 'visible_por_defecto': True},
            {'campo': 'fecha_registro', 'etiqueta': 'FECHA DE REGISTRO', 'visible_por_defecto': True},
            {'campo': 'nombre', 'etiqueta': 'NOMBRE COMPLETO', 'visible_por_defecto': True},
            {'campo': 'tipo_documento', 'etiqueta': 'TIPO DOCUMENTO', 'visible_por_defecto': True},
            {'campo': 'cedula', 'etiqueta': 'N° DOCUMENTO', 'visible_por_defecto': True},
            {'campo': 'genero', 'etiqueta': 'GÉNERO', 'visible_por_defecto': True},
            {'campo': 'telefono', 'etiqueta': 'TELÉFONO', 'visible_por_defecto': True},
            {'campo': 'email', 'etiqueta': 'CORREO ELECTRÓNICO', 'visible_por_defecto': True},
            {'campo': 'barrio', 'etiqueta': 'BARRIO', 'visible_por_defecto': True},
            {'campo': 'comuna', 'etiqueta': 'COMUNA', 'visible_por_defecto': True},
            {'campo': 'firma', 'etiqueta': 'FIRMA', 'visible_por_defecto': True}
        ]
        
        # Filtrar columnas según lo solicitado
        if columnas:
            # Filtrar solo las columnas que existen en columnas_disponibles
            columnas_seleccionadas = [col for col in columnas 
                                    if any(c['campo'] == col for c in columnas_disponibles)]
            
            # Si no hay columnas válidas, usar las visibles por defecto
            if not columnas_seleccionadas:
                columnas_seleccionadas = [c['campo'] for c in columnas_disponibles if c['visible_por_defecto']]
        else:
            # Si no se especifican columnas, usar las visibles por defecto
            columnas_seleccionadas = [c['campo'] for c in columnas_disponibles if c['visible_por_defecto']]
        
        # Obtener los encabezados para las columnas seleccionadas
        headers = []
        for col in columnas_disponibles:
            if col['campo'] in columnas_seleccionadas:
                headers.append(col['etiqueta'])
        
        # Agregar encabezados de la tabla
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[start_row].height = 30
        
        # Agregar datos de los asistentes
        if 'asistentes' in reunion and reunion['asistentes']:
            for row_num, asistente in enumerate(reunion['asistentes'], start=start_row + 1):
                for col_num, col_name in enumerate(columnas_seleccionadas, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    # Asignar el valor correspondiente
                    if col_name == 'numero':
                        cell.value = row_num - start_row
                    else:
                        cell.value = str(asistente.get(col_name, ''))
                    
                    # Aplicar estilos
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar el ancho de las columnas al contenido
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 30)
        
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            temp_filename = tmp.name
            
        # Guardar el libro de trabajo
        wb.save(temp_filename)
        
        # Leer el contenido del archivo
        with open(temp_filename, 'rb') as f:
            file_content = f.read()
        
        # Eliminar el archivo temporal
        os.unlink(temp_filename)
        
        # Crear respuesta
        from flask import make_response
        from datetime import datetime
        
        response = make_response(file_content)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=asistencia_reunion_{reunion_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        logger.error(f"Error al exportar reunión a Excel: {str(e)}", exc_info=True)
        return {
            'success': False,
            'message': 'Error al generar el archivo Excel',
            'error': str(e)
        }, 500
